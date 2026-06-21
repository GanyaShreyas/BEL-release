from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
import json
from pymongo import MongoClient, ReturnDocument
from bson import ObjectId
import traceback
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import secrets
import hashlib
import csv
from io import StringIO,BytesIO
import re
import zipfile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os
from io import StringIO
from datetime import datetime
from zoneinfo import ZoneInfo

client = MongoClient("mongodb://localhost:27017/")
db = client["inventory_db"]  # Use your DB name
collection = db["product_details"]
log_collection = db["api_logs"]
users_collection = db["users"]
sessions_collection = db["sessions"]
spares_master = db["spares_master"]   # master list
spares_in_col = db["spares_in"]       # logs table
spares_out_col = db["spares_out"]  # logs table
spares_out_returnable_col = db["spares_out_returnable"]
spares_in_returned_col = db["spares_in_returned"]
spares_returnable_requests_col = db["spares_returnable_requests"]
spares_audit = db["spares_audit"]
spares_stores_col = db["spares_stores"]  # { "name": str, ... }
obd_collection = db["obd_records"]
config_details_col = db["configuration_details"]

# Admin Projects collection
admin_projects_collection = db["admin_projects"]
# ----------------------
# Admin Projects Endpoints
# ----------------------

@csrf_exempt
def admin_add_project(request):
    """Add a new project (with duplicate check)"""
    user, err = require_auth(request, role="admin")
    if err:
        return err
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)
    try:
        body = json.loads(request.body or b"{}")
        project_name = (body.get("projectName") or "").strip()
        if not project_name:
            return JsonResponse({"error": "Project name required"}, status=400)
        exists = admin_projects_collection.find_one({"projectName": project_name})
        if exists:
            return JsonResponse({"error": "Project already exists"}, status=409)
        admin_projects_collection.insert_one({
            "projectName": project_name,
            "items": [],
            "createdBy": user.get("username"),
            "createdAt": datetime.now(ZoneInfo("Asia/Kolkata")),
        })
        return JsonResponse({"message": "Project created", "projectName": project_name}, status=201)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def admin_add_item(request):
    """Add item data to a project (with duplicate hierarchy check)"""
    user, err = require_auth(request, role="admin")
    if err:
        return err
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)
    try:
        body = json.loads(request.body or b"{}")
        project_name = (body.get("projectName") or "").strip()
        item_type = (body.get("itemType") or "").strip()
        item_name = (body.get("itemName") or "").strip()
        part_no = (body.get("partNo") or "").strip()
        if not (project_name and item_type and item_name and part_no):
            return JsonResponse({"error": "All fields required"}, status=400)
        # Check for duplicate in hierarchy
        exists = admin_projects_collection.find_one({
            "projectName": project_name,
            "items": {
                "$elemMatch": {
                    "partNo": part_no
                }
            }
        })
        if exists:
            return JsonResponse({"error": "Duplicate item in hierarchy"}, status=409)
        # Add item
        result = admin_projects_collection.update_one(
            {"projectName": project_name},
            {"$push": {"items": {
                "itemType": item_type,
                "itemName": item_name,
                "partNo": part_no,
                "createdBy": user.get("username"),
                "createdAt": datetime.now(ZoneInfo("Asia/Kolkata"))
            }}}
        )
        if result.matched_count == 0:
            return JsonResponse({"error": "Project not found"}, status=404)
        return JsonResponse({"message": "Item added"}, status=201)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def admin_edit_item(request):
    """Edit item data in a project"""
    user, err = require_auth(request, role="admin")
    if err:
        return err
    if request.method != "PUT":
        return JsonResponse({"error": "Only PUT allowed"}, status=405)
    try:
        body = json.loads(request.body or b"{}")
        project_name = (body.get("projectName") or "").strip()
        items = body.get("items")
        # If batch replace (Save All)
        if project_name and isinstance(items, list):
            result = admin_projects_collection.update_one(
                {"projectName": project_name},
                {"$set": {"items": items}}
            )
            if result.matched_count == 0:
                return JsonResponse({"error": "Project not found"}, status=404)
            return JsonResponse({"message": "All items replaced successfully", "count": len(items)})
        # Else, single item edit
        old_item_type = (body.get("oldItemType") or body.get("itemType") or "").strip()
        old_item_name = (body.get("oldItemName") or body.get("itemName") or "").strip()
        old_part_no = (body.get("oldPartNo") or body.get("partNo") or "").strip()
        new_data = body.get("newData") or {}
        if not (project_name and old_item_type and old_item_name and old_part_no and new_data):
            return JsonResponse({"error": "All fields required"}, status=400)
        result = admin_projects_collection.update_one(
            {"projectName": project_name, "items": {
                "$elemMatch": {
                    "itemType": old_item_type,
                    "itemName": old_item_name,
                    "partNo": old_part_no
                }
            }},
            {"$set": {"items.$": {
                **new_data,
                "updatedBy": user.get("username"),
                "updatedAt": datetime.now(ZoneInfo("Asia/Kolkata"))
            }}}
        )
        if result.matched_count == 0:
            return JsonResponse({"error": "Item not found"}, status=404)
        return JsonResponse({"message": "Item updated"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def admin_delete_item(request):
    """Delete item data from a project"""
    user, err = require_auth(request, role="admin")
    if err:
        return err
    if request.method != "DELETE":
        return JsonResponse({"error": "Only DELETE allowed"}, status=405)
    try:
        body = json.loads(request.body or b"{}")
        project_name = (body.get("projectName") or "").strip()
        item_type = (body.get("itemType") or "").strip()
        item_name = (body.get("itemName") or "").strip()
        part_no = (body.get("partNo") or "").strip()
        if not (project_name and item_type and item_name and part_no):
            return JsonResponse({"error": "All fields required"}, status=400)
        result = admin_projects_collection.update_one(
            {"projectName": project_name},
            {"$pull": {"items": {
                "itemType": item_type,
                "itemName": item_name,
                "partNo": part_no
            }}}
        )
        if result.matched_count == 0:
            return JsonResponse({"error": "Project not found"}, status=404)
        return JsonResponse({"message": "Item deleted"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def admin_get_projects(request):
    """Get all projects"""
    user, err = require_auth(request)
    if err:
        return err
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)
    try:
        projects = list(admin_projects_collection.find({}, {"_id": 0, "projectName": 1}))
        project_names = [p.get("projectName") for p in projects if p.get("projectName") is not None]
        project_names_sorted = sorted(project_names, key=lambda s: str(s).casefold())
        return JsonResponse({"projects": project_names_sorted})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def admin_get_project_items(request):
    """Get all items for a project"""
    user, err = require_auth(request)
    if err:
        return err
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)
    try:
        project_name = request.GET.get("projectName", "").strip()
        if not project_name:
            return JsonResponse({"error": "Project name required"}, status=400)
        doc = admin_projects_collection.find_one({"projectName": project_name}, {"_id": 0, "items": 1})
        if not doc:
            return JsonResponse({"error": "Project not found"}, status=404)
        return JsonResponse({"items": doc.get("items", [])})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def admin_backup_mongo(request):
    """
    Admin-only backup: export every MongoDB collection as JSON (one JSON per collection)
    and return them as a single ZIP download.
    """
    user, err = require_auth(request)
    if err:
        return err

    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    try:
        now = datetime.now(ZoneInfo("Asia/Kolkata"))
        date_part = now.strftime("%d-%m-%Y")

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for collection_name in db.list_collection_names():
                coll = db[collection_name]
                docs = list(coll.find({}))
                json_text = json.dumps(docs, default=str, ensure_ascii=False, indent=2)
                json_filename = f"{collection_name}_{date_part}.json"
                zf.writestr(json_filename, json_text)

        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="mongo_backup_{date_part}.zip"'
        return response
    except Exception as e:
        stack_trace = traceback.format_exc()
        return JsonResponse({"error": str(e), "stack_trace": stack_trace}, status=500)

# Bootstrap a default admin user if none exists
try:
    if users_collection.count_documents({}) == 0:
        users_collection.insert_one({
            "name": "Administrator",
            "username": "admin",
            "password_hash": hashlib.sha256(("bel_simple_salt" + "admin123").encode("utf-8")).hexdigest(),
            "role": "admin",
            "created_at": datetime.now(ZoneInfo("Asia/Kolkata"))
        })
except Exception:
    # If Mongo isn't reachable at import time, ignore; runtime endpoints will fail gracefully
    pass

# Clean up expired sessions (older than 24 hours)
def cleanup_expired_sessions():
    try:
        from datetime import timedelta
        cutoff_time = datetime.now(ZoneInfo("Asia/Kolkata")) - timedelta(hours=24)
        sessions_collection.delete_many({"created_at": {"$lt": cutoff_time}})
    except Exception:
        pass  # Ignore cleanup errors

# Logging function to store logs in MongoDB
def log_api_response(endpoint, method, request_data, response_data):
    log_entry = {
        "endpoint": endpoint,
        "method": method,
        "request_data": request_data,
        "response_data": response_data,
        "timestamp": datetime.now(ZoneInfo("Asia/Kolkata"))
    }
    log_collection.insert_one(log_entry)


# ----------------------
# Auth helpers
# ----------------------
def hash_password(password: str) -> str:
    salt = "bel_simple_salt"  # replace with env var in prod
    return hashlib.sha256((salt + password).encode("utf-8")).hexdigest()


def generate_token() -> str:
    return secrets.token_hex(32)


def get_auth_token_from_request(request):
    auth_header = request.headers.get("Authorization") or request.META.get("HTTP_AUTHORIZATION")
    if not auth_header:
        return None
    parts = auth_header.split()
    if len(parts) == 2 and parts[0].lower() == "bearer":
        return parts[1]
    return None


def get_user_from_token(request):
    token = get_auth_token_from_request(request)
    if not token:
        return None
    session = sessions_collection.find_one({"token": token})
    if not session:
        return None
    user = users_collection.find_one({"_id": session.get("user_id")})
    if not user:
        return None
    return {"id": str(user.get("_id")), "username": user.get("username"), "role": user.get("role"), "name": user.get("name")}


def require_auth(request, role: str | None = None):
    user = get_user_from_token(request)
    if not user:
        return None, JsonResponse({"error": "Unauthorized"}, status=401)
    if role and user.get("role") != role:
        return None, JsonResponse({"error": "Forbidden"}, status=403)
    return user, None


# ----------------------
# Auth endpoints
# ----------------------
@csrf_exempt
def login(request):
    if request.method != "POST":
        error_response = {"error": "Only POST allowed"}
        log_api_response("login", request.method, getattr(request, 'body', None), error_response)
        return JsonResponse(error_response, status=405)
    try:
        body = json.loads(request.body or b"{}")
        username = (body.get("username") or "").strip()
        password = body.get("password") or ""
        if not username or not password:
            response = {"error": "username and password are required"}
            log_api_response("login", request.method, body, response)
            return JsonResponse(response, status=400)

        user = users_collection.find_one({"username": username})
        if not user or user.get("password_hash") != hash_password(password):
            response = {"error": "Invalid credentials"}
            log_api_response("login", request.method, {"username": username}, response)
            return JsonResponse(response, status=401)

        # Clean up expired sessions before creating new one
        cleanup_expired_sessions()
        
        token = generate_token()
        sessions_collection.insert_one({
            "token": token,
            "user_id": user.get("_id"),
            "role": user.get("role"),
            "created_at": datetime.now(ZoneInfo("Asia/Kolkata"))
        })

        response = {
            "token": token, 
            "role": user.get("role"),
            "username": user.get("username"),
            "name": user.get("name")
        }
        log_api_response("login", request.method, {"username": username}, response)
        return JsonResponse(response)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("login", request.method, getattr(request, 'body', None), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)


@csrf_exempt
def validate_token(request):
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("validate_token", request.method, None, error_response)
        return JsonResponse(error_response, status=405)
    
    user, err = require_auth(request)
    if err:
        return err
    
    try:
        # Clean up expired sessions
        cleanup_expired_sessions()
        
        # If we get here, the token is valid
        response = {
            "valid": True,
            "username": user.get("username"),
            "role": user.get("role"),
            "name": user.get("name")
        }
        log_api_response("validate_token", request.method, None, response)
        return JsonResponse(response)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("validate_token", request.method, None, {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)


@csrf_exempt
def logout(request):
    if request.method != "POST":
        error_response = {"error": "Only POST allowed"}
        log_api_response("logout", request.method, None, error_response)
        return JsonResponse(error_response, status=405)
    
    try:
        token = get_auth_token_from_request(request)
        if token:
            # Remove the session from the database
            sessions_collection.delete_one({"token": token})
        
        response = {"message": "Logged out successfully"}
        log_api_response("logout", request.method, None, response)
        return JsonResponse(response)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("logout", request.method, None, {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)


@csrf_exempt
def admin_add_user(request):
    if request.method != "POST":
        error_response = {"error": "Only POST allowed"}
        log_api_response("admin_add_user", request.method, getattr(request, 'body', None), error_response)
        return JsonResponse(error_response, status=405)

    user, err = require_auth(request, role="admin")
    if err:
        return err
    try:
        body = json.loads(request.body or b"{}")
        name = (body.get("name") or "").strip()
        username = (body.get("username") or "").strip().lower()
        password = body.get("password") or ""
        role = (body.get("role") or "user").strip().lower()

        if not name or not username or not password:
            response = {"error": "name, username, password are required"}
            log_api_response("admin_add_user", request.method, body, response)
            return JsonResponse(response, status=400)

        if role not in ["admin", "user"]:
            role = "user"

        exists = users_collection.find_one({"username": username})
        if exists:
            response = {"error": "username already exists"}
            log_api_response("admin_add_user", request.method, {"username": username}, response)
            return JsonResponse(response, status=409)

        doc = {
            "name": name,
            "username": username,
            "password_hash": hash_password(password),
            "role": role,
            "created_at": datetime.now(ZoneInfo("Asia/Kolkata"))
        }
        users_collection.insert_one(doc)
        response = {"message": "user created", "username": username, "role": role}
        log_api_response("admin_add_user", request.method, {"admin": user.get("username"), "new_user": username}, response)
        return JsonResponse(response, status=201)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("admin_add_user", request.method, getattr(request, 'body', None), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)


def admin_list_users(request):
    """Admin-only: list all users (excluding password hashes)."""
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    user, err = require_auth(request, role="admin")
    if err:
        return err

    try:
        docs = list(users_collection.find({}, {"password_hash": 0}))
        users_list = []
        for d in docs:
            users_list.append({
                "id": str(d.get("_id")),
                "name": d.get("name", ""),
                "username": d.get("username", ""),
                "role": d.get("role", ""),
                "designation": d.get("designation", ""),
                "mobile": d.get("mobile", ""),
                "created_at": str(d.get("created_at", "")),
            })
        return JsonResponse({"users": users_list})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def admin_edit_user(request):
    """Admin-only: edit user details (name, designation, mobile, store)."""
    if request.method != "PUT":
        return JsonResponse({"error": "Only PUT allowed"}, status=405)

    user, err = require_auth(request, role="admin")
    if err:
        return err

    try:
        body = json.loads(request.body or b"{}")
        username = (body.get("username") or "").strip().lower()
        if not username:
            return JsonResponse({"error": "username is required"}, status=400)

        target = users_collection.find_one({"username": username})
        if not target:
            return JsonResponse({"error": "User not found"}, status=404)

        update_fields = {}
        if "name" in body:
            update_fields["name"] = (body["name"] or "").strip()
        if "designation" in body:
            update_fields["designation"] = (body["designation"] or "").strip()
        if "mobile" in body:
            update_fields["mobile"] = (body["mobile"] or "").strip()

        if not update_fields:
            return JsonResponse({"error": "No fields to update"}, status=400)

        users_collection.update_one({"username": username}, {"$set": update_fields})
        return JsonResponse({"message": f"User '{username}' updated successfully"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def admin_delete_user(request):
    """Admin-only: delete a user by username."""
    if request.method != "DELETE":
        return JsonResponse({"error": "Only DELETE allowed"}, status=405)

    user, err = require_auth(request, role="admin")
    if err:
        return err

    try:
        body = json.loads(request.body or b"{}")
        username = (body.get("username") or "").strip().lower()

        if not username:
            return JsonResponse({"error": "username is required"}, status=400)

        # Prevent admin from deleting themselves
        if username == user.get("username"):
            return JsonResponse({"error": "Cannot delete your own account"}, status=400)

        target = users_collection.find_one({"username": username})
        if not target:
            return JsonResponse({"error": "User not found"}, status=404)

        users_collection.delete_one({"username": username})
        # Also remove their sessions
        sessions_collection.delete_many({"user_id": target.get("_id")})

        return JsonResponse({"message": f"User '{username}' deleted successfully"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def admin_reset_password(request):
    """Admin-only: reset any user's password by username."""
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    user, err = require_auth(request, role="admin")
    if err:
        return err

    try:
        body = json.loads(request.body or b"{}")
        username = (body.get("username") or "").strip().lower()
        new_password = body.get("new_password") or ""

        if not username:
            return JsonResponse({"error": "username is required"}, status=400)
        if not new_password or len(new_password) < 4:
            return JsonResponse({"error": "new_password must be at least 4 characters"}, status=400)

        target = users_collection.find_one({"username": username})
        if not target:
            return JsonResponse({"error": "User not found"}, status=404)

        users_collection.update_one(
            {"username": username},
            {"$set": {"password_hash": hash_password(new_password)}}
        )
        return JsonResponse({"message": "Password reset successfully", "username": username})
    except Exception as e:
        stack = traceback.format_exc()
        return JsonResponse({"error": str(e), "stack_trace": stack}, status=500)


@csrf_exempt
def user_change_password(request):
    """Authenticated user changes their own password (current + new + confirm)."""
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        body = json.loads(request.body or b"{}")
        current_password = body.get("current_password") or ""
        new_password = body.get("new_password") or ""
        confirm_password = body.get("confirm_password") or ""

        if not current_password:
            return JsonResponse({"error": "current_password is required"}, status=400)
        if not new_password or len(new_password) < 4:
            return JsonResponse({"error": "new_password must be at least 4 characters"}, status=400)
        if new_password != confirm_password:
            return JsonResponse({"error": "new_password and confirm_password do not match"}, status=400)

        uid = user.get("id")
        doc = users_collection.find_one({"_id": ObjectId(uid)}) if uid else None
        if not doc:
            return JsonResponse({"error": "User not found"}, status=404)
        if doc.get("password_hash") != hash_password(current_password):
            return JsonResponse({"error": "Current password is incorrect"}, status=400)

        users_collection.update_one(
            {"_id": ObjectId(uid)},
            {"$set": {"password_hash": hash_password(new_password)}}
        )
        return JsonResponse({"message": "Password changed successfully"})
    except Exception as e:
        stack = traceback.format_exc()
        return JsonResponse({"error": str(e), "stack_trace": stack}, status=500)


# ----------------------
# Inventory: Item In / Out / CRUD by passNo
# ----------------------
@csrf_exempt
def items_in(request):
    if request.method != "POST":
        error_response = {"error": "Only POST allowed"}
        log_api_response("items_in", request.method, getattr(request, 'body', None), error_response)
        return JsonResponse(error_response, status=405)
    user, err = require_auth(request)
    if err:
        return err
    try:
        body = json.loads(request.body or b"{}")
        pass_no = (body.get("passNo") or "").strip()
        date_in = (body.get("dateIn") or datetime.now(ZoneInfo("Asia/Kolkata")).date().isoformat())
        customer = {
            "name": body.get("customerName"),
            "unitAddress": body.get("customerUnitAddress"),
            "location": body.get("customerLocation"),
            "phone": body.get("customerPhoneNo"),
        }
        project_name = body.get("projectName")
        items = body.get("items") or []
        year_raw = body.get("yearOfMfg")  # legacy: applied to items missing yearOfMfg

        if not pass_no:
            response = {"error": "passNo is required"}
            log_api_response("items_in", request.method, body, response)
            return JsonResponse(response, status=400)

        exists = collection.find_one({"passNo": pass_no})
        if exists:
            response = {"error": "passNo already exists"}
            log_api_response("items_in", request.method, {"passNo": pass_no}, response)
            return JsonResponse(response, status=409)

        legacy_pass_year = None
        if year_raw not in (None, ""):
            legacy_pass_year, yerr = _parse_optional_year_mfg_value(year_raw)
            if yerr:
                return yerr

        normalized_items = []
        for it in items:
            yr_raw = it.get("yearOfMfg")
            if yr_raw in (None, "") and legacy_pass_year is not None:
                yr_raw = legacy_pass_year
            y_item, yerr = _parse_optional_year_mfg_value(yr_raw)
            if yerr:
                return yerr

            row = {
                "equipmentType": it.get("equipmentType"),
                "itemName": it.get("itemName"),
                "partNumber": it.get("partNumber"),
                "serialNumber": it.get("serialNumber"),
                "defectDetails": it.get("defectDetails"),
                "itemIn": True,  # Always true when item is entered
                "itemOut": False,
                "dateOut": None,  # Will be set when item goes out
                "dispatchThrough": "",
                "itemRectificationDetails": "",  # New field for rectification details
                "itemFeedback1Details": "",  # New field for feedback 1 details
                "itemFeedback2Details": "",  # New field for feedback 2 details
            }
            if y_item is not None:
                row["yearOfMfg"] = y_item
            normalized_items.append(row)

        # Sort items by part number in ascending order
        normalized_items.sort(key=lambda x: x.get("partNumber", ""))

        doc = {
            "passNo": pass_no,
            "dateIn": date_in,
            "customer": customer,
            "projectName": project_name,
            "items": normalized_items,
            "createdBy": user.get("username"),
            "createdAt": datetime.now(ZoneInfo("Asia/Kolkata")),
            "updatedAt": datetime.now(ZoneInfo("Asia/Kolkata")),
        }
        collection.insert_one(doc)
        doc.pop("_id", None)
        response = {"message": "Item In recorded", "data": doc}
        log_api_response("items_in", request.method, {"passNo": pass_no}, response)
        return JsonResponse(response, status=201)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("items_in", request.method, getattr(request, 'body', None), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

def get_item_by_passno(request, pass_no):
    if not pass_no:
        print("Enter a pass number")  # <--- your print statement
        error_response = {"error": "passNo cannot be empty"}
        log_api_response("get_item_by_passno", request.method, {"passNo": pass_no}, error_response)
        return JsonResponse(error_response, status=400)
    
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("get_item_by_passno", request.method, {"passNo": pass_no}, error_response)
        return JsonResponse(error_response, status=405)
    user, err = require_auth(request)
    if err:
        return err
    try:
        doc = collection.find_one({"passNo": pass_no}, {"_id": 0})
        if not doc:
            response = {"error": "Not found"}
            log_api_response("get_item_by_passno", request.method, {"passNo": pass_no}, response)
            return JsonResponse(response, status=404)

        # Normalize legacy fields on read (RFC/RFD + Remarks 2 compatibility)
        _normalize_items_on_read(doc.get("items", []))
        _attach_legacy_year_mfg(doc)

        log_api_response("get_item_by_passno", request.method, {"passNo": pass_no}, doc)
        return JsonResponse(doc, safe=False)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("get_item_by_passno", request.method, {"passNo": pass_no}, {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

@csrf_exempt
def update_item_rfd(request, pass_no):
    if request.method != "PUT":
        error_response = {"error": "Only PUT allowed"}
        log_api_response("update_item_rfd", request.method, {"passNo": pass_no}, error_response)
        return JsonResponse(error_response, status=405)
    user, err = require_auth(request)
    if err:
        return err
    try:
        print(f"=== ITEM RFD UPDATE DEBUG ===")
        print(f"Pass Number: {pass_no}")
        print(f"User: {user.get('username')}")
        print(f"Request body: {request.body}")

        body = json.loads(request.body or b"{}")
        updates = body.get("items") or []

        print(f"Parsed body: {body}")
        print(f"Updates array: {updates}")

        doc = collection.find_one({"passNo": pass_no})
        if not doc:
            response = {"error": "Not found"}
            log_api_response("update_item_rfd", request.method, {"passNo": pass_no}, response)
            return JsonResponse(response, status=404)
        
        print(f"Found document: {doc.get('passNo')}")
        print(f"Original items: {doc.get('items')}")

        # Check if we have the same number of items
        original_items = doc.get("items", [])
        if len(updates) != len(original_items):
            response = {"error": f"Number of items mismatch. Expected {len(original_items)}, got {len(updates)}"}
            return JsonResponse(response, status=400)

        print(f"DEBUG: Received updates: {updates}")

        # Update items by position (index) instead of serial number
        new_items = []
        for i, (original_item, update_item) in enumerate(zip(original_items, updates)):
            print(f"Processing item {i}: original={original_item.get('serialNumber')}, update={update_item.get('serialNumber')}")

            # Create updated item
            updated_item = original_item.copy()
            updated_item["itemRfd"] = bool(update_item.get("itemRfd", False))

            # Prevent RFD being unset for itemOut items
            if original_item.get("itemOut") is True:
                updated_item["itemRfd"] = True

            # Handle dateRfd
            if updated_item["itemRfd"]:
                if update_item.get("dateRfd"):
                    updated_item["dateRfd"] = update_item["dateRfd"]
                elif not original_item.get("dateRfd"):
                    updated_item["dateRfd"] = datetime.now(ZoneInfo("Asia/Kolkata")).date().isoformat()
                    print(f"DEBUG: Auto-setting dateRfd for item {i} to {updated_item['dateRfd']}")
            else:
                updated_item["dateRfd"] = None
                print(f"DEBUG: Clearing dateRfd for item {i} since itemRfd is False")

            # Handle rectification details
            if "itemRectificationDetails" in update_item:
                updated_item["itemRectificationDetails"] = update_item["itemRectificationDetails"] or ""

            if "itemFeedback1Details" in update_item:
                updated_item["itemFeedback1Details"] = update_item["itemFeedback1Details"] or ""

            if "itemFeedback2Details" in update_item:
                updated_item["itemFeedback2Details"] = update_item["itemFeedback2Details"] or ""
            
            print(f"Updated item {i}: {updated_item}")
            new_items.append(updated_item)    

        print(f"Final items array: {new_items}")

        result = collection.update_one({"passNo": pass_no}, {"$set": {"items": new_items, "updatedAt": datetime.now(ZoneInfo("Asia/Kolkata")), "updatedBy": user.get("username")}})
        print(f"Update result: matched={result.matched_count}, modified={result.modified_count}")

        response = {"message": "RFD statuses updated"}
        log_api_response("update_item_rfd", request.method, {"passNo": pass_no, "updates_count": len(updates)}, response)
        return JsonResponse(response)
    except Exception as e:
        print(f"ERROR in update_item_rfd: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("update_item_rfd", request.method, {"passNo": pass_no}, {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

@csrf_exempt
def update_item_out(request, pass_no):
    if request.method != "PUT":
        error_response = {"error": "Only PUT allowed"}
        log_api_response("update_item_out", request.method, {"passNo": pass_no}, error_response)
        return JsonResponse(error_response, status=405)
    user, err = require_auth(request)
    if err:
        return err
    try:
        print(f"=== ITEM OUT UPDATE DEBUG ===")
        print(f"Pass Number: {pass_no}")
        print(f"User: {user.get('username')}")
        print(f"Request body: {request.body}")
        
        body = json.loads(request.body or b"{}")
        updates = body.get("items") or []

        print(f"Parsed body: {body}")
        print(f"Updates array: {updates}")

        doc = collection.find_one({"passNo": pass_no})
        if not doc:
            response = {"error": "Not found"}
            log_api_response("update_item_out", request.method, {"passNo": pass_no}, response)
            return JsonResponse(response, status=404)

        print(f"Found document: {doc.get('passNo')}")
        print(f"Original items: {doc.get('items')}")

        # Check if we have the same number of items
        original_items = doc.get("items", [])
        # Normalize legacy fields (itemRfc -> itemRfd) so guards work correctly
        _normalize_items_on_read(original_items)
        if len(updates) != len(original_items):
            response = {"error": f"Number of items mismatch. Expected {len(original_items)}, got {len(updates)}"}
            return JsonResponse(response, status=400)

        print(f"DEBUG: Received updates: {updates}")
        
        # Update items by position (index) instead of serial number
        new_items = []
        for i, (original_item, update_item) in enumerate(zip(original_items, updates)):
            print(f"Processing item {i}: original={original_item.get('serialNumber')}, update={update_item.get('serialNumber')}")
            
            # Create updated item
            updated_item = original_item.copy()
            updated_item["itemOut"] = bool(update_item.get("itemOut", False))
            
            if not original_item.get("itemRfd"):
                updated_item["itemOut"] = False

            if original_item.get("itemOut") is True:
                updated_item["itemOut"] = True

            # Handle dateOut
            if updated_item["itemOut"]:
                if update_item.get("dateOut"):
                    updated_item["dateOut"] = update_item["dateOut"]
                elif not original_item.get("dateOut"):
                    updated_item["dateOut"] = datetime.now(ZoneInfo("Asia/Kolkata")).date().isoformat()
                    print(f"DEBUG: Auto-setting dateOut for item {i} to {updated_item['dateOut']}")

                # Legacy handling: rows that were already Item Out before this update
                # may not have dispatchThrough populated; allow null/blank for them.
                if original_item.get("itemOut") is True:
                    updated_item["dispatchThrough"] = update_item.get("dispatchThrough")
                else:
                    dispatch_through = (update_item.get("dispatchThrough") or "").strip()
                    if dispatch_through not in ("Direct Collection", "Through Shipping"):
                        return JsonResponse(
                            {"error": "Dispatch Through is required for all newly Item Out rows"},
                            status=400,
                        )
                    updated_item["dispatchThrough"] = dispatch_through
            else:
                updated_item["dateOut"] = None
                updated_item["dispatchThrough"] = ""
                print(f"DEBUG: Clearing dateOut for item {i} since itemOut is False")
            
            # Handle rectification details
            if "itemRectificationDetails" in update_item:
                updated_item["itemRectificationDetails"] = update_item["itemRectificationDetails"] or ""

            if "itemFeedback1Details" in update_item:
                updated_item["itemFeedback1Details"] = update_item["itemFeedback1Details"] or ""

            if "itemFeedback2Details" in update_item:
                updated_item["itemFeedback2Details"] = update_item["itemFeedback2Details"] or ""
            
            print(f"Updated item {i}: {updated_item}")
            new_items.append(updated_item)

        print(f"Final items array: {new_items}")
        
        result = collection.update_one({"passNo": pass_no}, {"$set": {"items": new_items, "updatedAt": datetime.now(ZoneInfo("Asia/Kolkata")), "updatedBy": user.get("username")}})
        print(f"Update result: matched={result.matched_count}, modified={result.modified_count}")
        
        response = {"message": "ItemOut statuses updated"}
        log_api_response("update_item_out", request.method, {"passNo": pass_no, "updates_count": len(updates)}, response)
        return JsonResponse(response)
    except Exception as e:
        print(f"ERROR in update_item_out: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("update_item_out", request.method, {"passNo": pass_no}, {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

@csrf_exempt
def edit_record(request, pass_no):
    user, err = require_auth(request)
    if err:
        return err
    try:
        if request.method == "GET":
            doc = collection.find_one({"passNo": pass_no}, {"_id": 0})
            if not doc:
                response = {"error": "Entry Not found"}
                log_api_response("edit_record", request.method, {"passNo": pass_no}, response)
                return JsonResponse(response, status=404)

            # Normalize legacy fields on read (RFC/RFD + Remarks 2 compatibility)
            _normalize_items_on_read(doc.get("items", []))
            _attach_legacy_year_mfg(doc)

            log_api_response("edit_record", request.method, {"passNo": pass_no}, doc)
            return JsonResponse(doc, safe=False)
        elif request.method == "PUT":
            body = json.loads(request.body or b"{}")
            if body.get("passNo") and body.get("passNo") != pass_no:
                response = {"error": "passNo cannot be changed"}
                log_api_response("edit_record", request.method, {"passNo": pass_no}, response)
                return JsonResponse(response, status=400)

            existing_doc = collection.find_one({"passNo": pass_no}, {"items": 1})
            if not existing_doc:
                response = {"error": "Not found"}
                log_api_response("edit_record", request.method, {"passNo": pass_no}, response)
                return JsonResponse(response, status=404)

            existing_items = existing_doc.get("items", [])
            already_out_keys = set()
            for it in existing_items:
                if it.get("itemOut"):
                    already_out_keys.add((
                        str(it.get("equipmentType") or ""),
                        str(it.get("itemName") or ""),
                        str(it.get("partNumber") or ""),
                        str(it.get("serialNumber") or ""),
                    ))

            allowed_fields = ["dateIn", "customer", "projectName", "items"]
            set_fields = {k: v for k, v in body.items() if k in allowed_fields}

            # Sort items by part number if items are being updated
            if "items" in set_fields:
                for item in set_fields["items"]:
                    item["itemIn"] = True  # Always true when item is entered
                    if "dateOut" not in item:
                        item["dateOut"] = None
                    if "dispatchThrough" not in item:
                        item["dispatchThrough"] = ""
                    dispatch_through = (item.get("dispatchThrough") or "").strip()
                    item_key = (
                        str(item.get("equipmentType") or ""),
                        str(item.get("itemName") or ""),
                        str(item.get("partNumber") or ""),
                        str(item.get("serialNumber") or ""),
                    )
                    was_already_out = item_key in already_out_keys
                    if item.get("itemOut"):
                        if (not was_already_out) and dispatch_through not in ("Direct Collection", "Through Shipping"):
                            return JsonResponse(
                                {"error": "Dispatch Through is required for all newly Item Out rows"},
                                status=400,
                            )
                        if was_already_out:
                            item["dispatchThrough"] = item.get("dispatchThrough")
                        else:
                            item["dispatchThrough"] = dispatch_through
                    else:
                        item["dispatchThrough"] = ""
                    if "dateRfd" not in item:
                        item["dateRfd"] = None
                    if "itemRectificationDetails" not in item:
                        item["itemRectificationDetails"] = ""
                    if "itemFeedback1Details" not in item:
                        item["itemFeedback1Details"] = ""
                    if "itemFeedback2Details" not in item:
                        item["itemFeedback2Details"] = ""
                    raw_y = item.get("yearOfMfg")
                    if raw_y in (None, ""):
                        item.pop("yearOfMfg", None)
                    else:
                        y, yerr = _parse_optional_year_mfg_value(raw_y)
                        if yerr:
                            return yerr
                        item["yearOfMfg"] = y
                set_fields["items"].sort(key=lambda x: x.get("partNumber", ""))
            
            set_fields["updatedAt"] = datetime.now(ZoneInfo("Asia/Kolkata"))
            set_fields["updatedBy"] = user.get("username")
            result = collection.update_one({"passNo": pass_no}, {"$set": set_fields})
            if result.matched_count == 0:
                response = {"error": "Not found"}
                log_api_response("edit_record", request.method, {"passNo": pass_no}, response)
                return JsonResponse(response, status=404)
            response = {"message": "Record updated"}
            log_api_response("edit_record", request.method, {"passNo": pass_no}, response)
            return JsonResponse(response)
        elif request.method == "DELETE":
            result = collection.delete_one({"passNo": pass_no})
            if result.deleted_count == 0:
                response = {"error": "Not found"}
                log_api_response("edit_record", request.method, {"passNo": pass_no}, response)
                return JsonResponse(response, status=404)
            response = {"message": "Record deleted"}
            log_api_response("edit_record", request.method, {"passNo": pass_no}, response)
            return JsonResponse(response)
        else:
            error_response = {"error": "Method not allowed"}
            log_api_response("edit_record", request.method, {"passNo": pass_no}, error_response)
            return JsonResponse(error_response, status=405)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("edit_record", request.method, {"passNo": pass_no}, {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)


# ----------------------
# Search + download
# ----------------------

def _normalize_item_fields_on_read(item: dict):
    """
    Normalize legacy DB keys without modifying stored documents.
    - RFC -> RFD compatibility
    - remarks_2 / "Handed Over To / Dispatched Details" -> itemFeedback2Details
    """
    if not isinstance(item, dict):
        return item

    # RFC -> RFD compatibility (read-time normalization only)
    # If either flag is set, treat it as RFD for UI/reporting purposes.
    rfd_status = item.get("itemRfd", None)
    rfc_status = item.get("itemRfc", None)
    if rfc_status in (None, ""):
        rfc_status = item.get("itemRFC", None)

    if rfd_status in (None, "", False) and rfc_status not in (None, "", False):
        item["itemRfd"] = rfc_status

    rfd_date = item.get("dateRfd", None)
    rfc_date = item.get("dateRfc", None)
    if rfc_date in (None, ""):
        rfc_date = item.get("dateRFC", None)

    if rfd_date in (None, "", False) and rfc_date not in (None, "", False):
        item["dateRfd"] = rfc_date

    if "dispatchThrough" not in item or item.get("dispatchThrough") is None:
        item["dispatchThrough"] = ""

    # Remarks 2 compatibility (read-time normalization only)
    if ("itemFeedback2Details" not in item) or (item.get("itemFeedback2Details") in (None, "")):
        feedback2 = item.get("remarks_2")
        if feedback2 in (None, ""):
            feedback2 = item.get("Handed Over To / Dispatched Details")
        if feedback2 is not None:
            item["itemFeedback2Details"] = feedback2

    return item


def _normalize_items_on_read(items):
    if not isinstance(items, list):
        return items
    for it in items:
        _normalize_item_fields_on_read(it)
    return items


def _attach_legacy_year_mfg(doc):
    """Attach doc-level yearOfMfg to items missing per-item value (read-time only)."""
    if not isinstance(doc, dict):
        return
    legacy = doc.get("yearOfMfg")
    if legacy in (None, "", False):
        return
    items = doc.get("items")
    if not isinstance(items, list):
        return
    for it in items:
        if not isinstance(it, dict):
            continue
        if it.get("yearOfMfg") in (None, "", False):
            it["yearOfMfg"] = legacy


def _parse_optional_year_mfg_value(raw):
    """
    Parse optional year for a single item.
    Returns (year_int_or_none, error_response_or_none).
    """
    if raw in (None, ""):
        return None, None
    try:
        y = int(str(raw).strip())
    except (TypeError, ValueError):
        return None, JsonResponse(
            {"error": "Year of MFG must be a whole number between 2000 and 2100"},
            status=400,
        )
    if y < 2000 or y > 2100:
        return None, JsonResponse(
            {"error": "Year of MFG must be between 2000 and 2100"},
            status=400,
        )
    return y, None


def _format_date_ddmmyyyy(date_value):
    """
    UI-only formatting helper for generated Excel downloads.
    Does not alter DB storage format.
    """
    if not date_value:
        return ""
    try:
        if isinstance(date_value, str):
            # Common ISO format: YYYY-MM-DD...
            if re.match(r"^\d{4}-\d{2}-\d{2}", date_value):
                base = date_value[:10]
                yyyy, mm, dd = base.split("-")
                return f"{dd}-{mm}-{yyyy}"
            # Already dd-mm-yyyy
            if re.match(r"^\d{2}-\d{2}-\d{4}$", date_value.strip()):
                return date_value.strip()
            return date_value

        if isinstance(date_value, datetime):
            return date_value.strftime("%d-%m-%Y")
    except Exception:
        return str(date_value)
    return str(date_value)


def _build_date_filter(from_str: str | None, to_str: str | None):
    if not from_str and not to_str:
        return None
    cond = {}
    if from_str:
        cond["$gte"] = from_str
    if to_str:
        cond["$lte"] = to_str
    return cond


def _build_search_query(params):
    search_type = params.get("type")
    value = params.get("value")
    dispatch_through = (params.get("dispatchThrough") or "All").strip()
    part_no_filter = (params.get("partNo") or params.get("part_no") or "").strip()
    from_date = params.get("from")
    to_date = params.get("to")
    serialProjectName = params.get("serialProjectName")
    query = {}
    if search_type == "passNo" and value:
        query["passNo"] = value
    elif search_type == "serialNumber" and value:
        if serialProjectName:
            query["projectName"] = {"$regex": re.escape(serialProjectName), "$options": "i"}    
        query["items.serialNumber"] = {"$regex": re.escape(value.upper()), "$options": "i"}
    elif search_type == "ItemPartNo" and value:
        query["items.partNumber"] = value
    elif search_type == "ProjectName" and value:
        query["projectName"] = value
        if part_no_filter:
            # Optional additional filter when searching by Project Name
            query["items.partNumber"] = part_no_filter
    elif search_type == "PhoneNumber" and value:
        # Phone numbers may be stored as strings or numbers; try prefix regex first.
        phone_digits = re.sub(r"\D+", "", str(value))
        if phone_digits:
            query["customer.phone"] = {"$regex": f"^{phone_digits}"}
    elif search_type == "DateRange":
        pass  # only date filter
    date_cond = _build_date_filter(from_date, to_date)
    if date_cond:
        # Use $elemMatch for array fields so a SINGLE item must satisfy the
        # full range, and exclude null values that would falsely match $lte.
        item_date_cond = {**date_cond, "$ne": None}
        query["$or"] = [
            {"dateIn": date_cond},
            {"items": {"$elemMatch": {"dateOut": item_date_cond}}},
            {"items": {"$elemMatch": {"dateRfd": item_date_cond}}},
            {"items": {"$elemMatch": {"dateRfc": item_date_cond}}},
        ]
    if dispatch_through and dispatch_through != "All":
        query["items.dispatchThrough"] = dispatch_through
    print(query)
    return query

def _filter_items_by_date_range(items, from_date, to_date):
    """Filter items to only those where dateOut, dateRfd, or dateRfc falls in range."""
    if not from_date and not to_date:
        return items
    filtered = []
    for item in items:
        for key in ("dateOut", "dateRfd", "dateRfc"):
            val = item.get(key)
            if not val:
                continue
            if from_date and val < from_date:
                continue
            if to_date and val > to_date:
                continue
            filtered.append(item)
            break
    return filtered


def _filter_items_by_dispatch_through(items, dispatch_through):
    if not dispatch_through or dispatch_through == "All":
        return items
    return [it for it in items if (it.get("dispatchThrough") or "") == dispatch_through]


def _filter_serial(items, serial_substring=None, status=None):
    """Filter items by serial number substring (case-insensitive) and status."""
    if not serial_substring:
        return items

    serial_substring = serial_substring.upper()
    filtered = []

    for item in items:
        serial = item.get("serialNumber", "")

        # 🔹 Serial number substring match
        if not (serial and serial_substring in serial.upper()):
            continue

        # 🔹 Status filter (optional)
        if status == "In" and not (item.get("itemIn") and not item.get("itemOut") and not item.get("itemRfd")):
            continue
        if status == "RFD" and not (item.get("itemIn") and item.get("itemRfd") and not item.get("itemOut")):
            continue
        if status == "Out" and not (item.get("itemIn") and item.get("itemOut")):
            continue

        filtered.append(item)

    return filtered

def _filter_items(items, part_number=None, status=None):
    """Filter items by part number and/or status (In/Out)."""
    filtered = []
    for item in items:
        # Part number filter
        if part_number and item.get("partNumber") != part_number:
            continue

        # Status filter
        if status == "In" and not (item.get("itemIn") and not item.get("itemOut") and not item.get("itemRfd")):
            continue

        if status == "RFD" and not (item.get("itemIn") and item.get("itemRfd") and not item.get("itemOut")):
            continue

        if status == "Out" and not (item.get("itemIn") and item.get("itemOut")):
            continue

        filtered.append(item)
    return filtered


def _filter_items_by_part_number(items, part_number):
    """Filter items to only include those matching the part number search"""
    if not part_number:
        return items
    return [item for item in items if item.get("partNumber") == part_number]


def _shape_search_result(doc):
    return {
        "passNo": doc.get("passNo"),
        "projectName": doc.get("projectName"),
        "dateIn": doc.get("dateIn"),
        "customer": doc.get("customer", {}),
        "yearOfMfg": doc.get("yearOfMfg"),
        "items": doc.get("items", []),
        "createdBy": doc.get("createdBy", ""),
        "updatedBy": doc.get("updatedBy", ""),
    }

def search(request):
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("search", request.method, dict(request.GET), error_response)
        return JsonResponse(error_response, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        params = request.GET
        query = _build_search_query(params)
        docs = list(collection.find(query))

        results = []
        serial_no = 1
        search_type = params.get("type")
        search_value = params.get("value")
        status = params.get("status")  # "In" or "Out"
        dispatch_through = (params.get("dispatchThrough") or "All").strip()
        part_no_filter = (params.get("partNo") or params.get("part_no") or "").strip()
        from_date = params.get("from")
        to_date = params.get("to")

        for doc in docs:
            filtered_items = doc.get("items", [])

            # Normalize legacy fields (RFC/RFD + Remarks 2 compatibility)
            _normalize_items_on_read(filtered_items)

            # Phone number matching fallback (covers numeric stored phone fields)
            if search_type == "PhoneNumber" and search_value:
                phone_digits = re.sub(r"\D+", "", str(search_value))
                doc_phone_digits = re.sub(r"\D+", "", str(doc.get("customer", {}).get("phone", "") or ""))
                if phone_digits and not re.match(rf"^{re.escape(phone_digits)}", doc_phone_digits):
                    continue

            if search_type == "serialNumber" and search_value:
                filtered_items = _filter_serial(filtered_items, serial_substring=search_value, status=status)
            elif search_type == "ItemPartNo" and search_value:
                filtered_items = _filter_items(filtered_items, part_number=search_value, status=status)
            elif search_type == "ProjectName" and part_no_filter:
                filtered_items = _filter_items(filtered_items, part_number=part_no_filter, status=status)
            elif status in ("In", "RFD", "Out"):
                filtered_items = _filter_items(filtered_items, status=status)

            filtered_items = _filter_items_by_dispatch_through(filtered_items, dispatch_through)

            # For date range: if the record didn't match on dateIn, filter items
            # to only those with dateOut/dateRfd/dateRfc in the range.
            if (from_date or to_date):
                date_in = doc.get("dateIn")
                date_in_matches = True
                if from_date and (not date_in or date_in < from_date):
                    date_in_matches = False
                if to_date and (not date_in or date_in > to_date):
                    date_in_matches = False
                if not date_in_matches:
                    filtered_items = _filter_items_by_date_range(filtered_items, from_date, to_date)

            for item in filtered_items:
                item["serialNo"] = serial_no
                serial_no += 1

            doc = {**doc, "items": filtered_items}
            _attach_legacy_year_mfg(doc)
            results.append(_shape_search_result({**doc, "_id": None}))

        response = {"count": len(results), "data": results}
        log_api_response("search", request.method, dict(params), {"count": response["count"]})
        return JsonResponse(response)

    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("search", request.method, dict(request.GET), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

@csrf_exempt
def search_download(request):
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("search_download", request.method, dict(request.GET), error_response)
        return JsonResponse(error_response, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        serial_no = 1
        params = request.GET
        query = _build_search_query(params)
        docs = list(collection.find(query))

        # Create CSV content
        output = StringIO()
        writer = csv.writer(output)

        # Column selection (optional). If absent, preserve existing export columns.
        columns_param = (params.get("columns") or "").strip()

        default_column_ids = [
            "slNo",
            "passNo",
            "projectName",
            "customerName",
            "customerUnitAddress",
            "customerLocation",
            "customerPhone",
            "equipmentType",
            "itemName",
            "partNumber",
            "serialNumber",
            "yearOfMfg",
            "defectDetails",
            "status",
            "dateIn",
            "dateRfd",
            "dateOut",
            "dispatchThrough",
            "rectificationDetails",
            "remarks2",
            "remarks1",
            "createdBy",
            "updatedBy",
        ]

        default_column_labels = {
            "slNo": "Sl No.",
            "passNo": "Pass No",
            "projectName": "Project Name",
            "customerName": "Customer Name",
            "customerUnitAddress": "Customer Unit Address",
            "customerLocation": "Customer Location",
            "customerPhone": "Customer Phone",
            "yearOfMfg": "Year of MFG",
            "equipmentType": "Equipment Type",
            "itemName": "Item Name",
            "partNumber": "Part Number",
            "serialNumber": "Serial Number",
            "defectDetails": "Defect Details",
            "status": "Status",
            "dateIn": "Date In",
            "dateRfd": "Date RFD",
            "dateOut": "Date Out",
            "dispatchThrough": "Dispatch Through",
            "rectificationDetails": "Item Rectification Details",
            "remarks2": "Feedback 2 details",
            "remarks1": "Feedback 1 details",
            "createdBy": "CreatedBy",
            "updatedBy": "updatedBy",
        }

        selected_column_labels = {
            "slNo": "SL NO.",
            "passNo": "PASS NO",
            "projectName": "PROJECT NAME",
            "customerName": "CUSTOMER NAME",
            "customerUnitAddress": "CUSTOMER UNIT ADDRESS",
            "customerLocation": "CUSTOMER LOCATION",
            "customerPhone": "CUSTOMER PHONE",
            "yearOfMfg": "YEAR OF MFG",
            "equipmentType": "EQUIPMENT TYPE",
            "itemName": "ITEM NAME",
            "partNumber": "PART NUMBER",
            "serialNumber": "SERIAL NUMBER",
            "defectDetails": "DEFECT DETAILS",
            "status": "STATUS",
            "dateIn": "DATE IN",
            "dateRfd": "DATE RFD",
            "dateOut": "DATE OUT",
            "dispatchThrough": "DISPATCH THROUGH",
            "rectificationDetails": "Rectification Details",
            "remarks2": "Handed Over To / Dispatched Details",
            "remarks1": "Remarks 1",
            "createdBy": "Created By",
            "updatedBy": "Updated By",
        }

        if columns_param:
            requested_ids = [c.strip() for c in columns_param.split(",") if c.strip()]
            allowed = set(selected_column_labels.keys())
            column_ids = [cid for cid in requested_ids if cid in allowed]
            if not column_ids:
                column_ids = default_column_ids
                column_labels = default_column_labels
            else:
                column_labels = selected_column_labels
        else:
            column_ids = default_column_ids
            column_labels = default_column_labels

        # Write header row
        writer.writerow([column_labels.get(cid, cid) for cid in column_ids])
        
        # Write data rows - one row per item
        for doc in docs:
            pass_no = doc.get("passNo", "")
            date_in = doc.get("dateIn", "")
            project_name = doc.get("projectName", "")
            customer = doc.get("customer", {})
            items = doc.get("items", [])
            createdBy = doc.get("createdBy", "")
            updatedBy = doc.get("updatedBy", "")
            
            search_type = params.get("type")
            status = params.get("status")
            dispatch_through = (params.get("dispatchThrough") or "All").strip()
            search_value = params.get("value")
            part_no_filter = (params.get("partNo") or params.get("part_no") or "").strip()

            # Normalize legacy fields (RFC/RFD + Remarks 2 compatibility)
            _normalize_items_on_read(items)

            # Phone number matching fallback (covers numeric stored phone fields)
            if search_type == "PhoneNumber" and search_value:
                phone_digits = re.sub(r"\D+", "", str(search_value))
                doc_phone_digits = re.sub(r"\D+", "", str(doc.get("customer", {}).get("phone", "") or ""))
                if phone_digits and not re.match(rf"^{re.escape(phone_digits)}", doc_phone_digits):
                    continue
            if search_type == "serialNumber" and search_value:
                items = _filter_serial(items, serial_substring=search_value, status=status)
            elif search_type == "ItemPartNo" and search_value:
                items = _filter_items(items, part_number = search_value, status=status)
            elif search_type == "ProjectName" and part_no_filter:
                items = _filter_items(items, part_number=part_no_filter, status=status)
            elif status in ("In", "RFD", "Out"):
                items = _filter_items(items, status=status)

            items = _filter_items_by_dispatch_through(items, dispatch_through)

            # For date range: if the record didn't match on dateIn, filter items
            # to only those with dateOut/dateRfd/dateRfc in the range.
            from_date = params.get("from")
            to_date = params.get("to")
            if (from_date or to_date):
                doc_date_in = doc.get("dateIn")
                date_in_matches = True
                if from_date and (not doc_date_in or doc_date_in < from_date):
                    date_in_matches = False
                if to_date and (not doc_date_in or doc_date_in > to_date):
                    date_in_matches = False
                if not date_in_matches:
                    items = _filter_items_by_date_range(items, from_date, to_date)

            # Filter items by part number if searching by part number
            
            # if search_type == "ItemPartNo" and params.get("value"):
            #     items = _filter_items_by_part_number(items, params.get("value"))

            ym_doc = doc.get("yearOfMfg")

            for item in items:
                # Determine status: OUT if both itemIn and itemOut are true, else IN
                status = "OUT" if item.get("itemIn") and item.get("itemOut") else "IN"
                status = "RFD" if item.get("itemIn") and item.get("itemRfd") and not item.get("itemOut") else status
                
                # Format phone number properly (remove scientific notation)
                phone = customer.get("phone", "")
                if phone and str(phone).isdigit():
                    phone = str(phone)
                
                # Format date properly for Excel
                date_rfd = item.get("dateRfd", "")
                if date_rfd:
                    # Ensure date is in YYYY-MM-DD format
                    try:
                        if isinstance(date_rfd, str):
                            date_rfd = date_rfd[:10]  # Take first 10 characters
                    except:
                        date_rfd = ""

                date_out = item.get("dateOut", "")
                if date_out:
                    # Ensure date is in YYYY-MM-DD format
                    try:
                        if isinstance(date_out, str):
                            date_out = date_out[:10]  # Take first 10 characters
                    except:
                        date_out = ""

                ym_item = item.get("yearOfMfg")
                if ym_item not in (None, ""):
                    year_mfg_cell = str(ym_item)
                elif ym_doc not in (None, ""):
                    year_mfg_cell = str(ym_doc)
                else:
                    year_mfg_cell = ""

                value_map = {
                    "slNo": serial_no,
                    "passNo": pass_no,
                    "projectName": project_name,
                    "customerName": customer.get("name", ""),
                    "customerUnitAddress": customer.get("unitAddress", ""),
                    "customerLocation": customer.get("location", ""),
                    "customerPhone": phone,
                    "yearOfMfg": year_mfg_cell,
                    "equipmentType": item.get("equipmentType", ""),
                    "itemName": item.get("itemName", ""),
                    "partNumber": item.get("partNumber", ""),
                    "serialNumber": item.get("serialNumber", ""),
                    "defectDetails": item.get("defectDetails", ""),
                    "status": status,
                    "dateIn": date_in,
                    "dateRfd": date_rfd,
                    "dateOut": date_out,
                    "dispatchThrough": item.get("dispatchThrough", ""),
                    "rectificationDetails": item.get("itemRectificationDetails", ""),
                    "remarks2": item.get("itemFeedback2Details", ""),
                    "remarks1": item.get("itemFeedback1Details", ""),
                    "createdBy": createdBy,
                    "updatedBy": updatedBy,
                }

                def _csv_safe(v):
                    s = str(v) if v is not None else ""
                    return s.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
                writer.writerow([_csv_safe(value_map.get(cid, "")) for cid in column_ids])
                serial_no += 1

        csv_content = output.getvalue()
        output.close()
        
        now = datetime.now(ZoneInfo('Asia/Kolkata'))
        default_filename = f"report_{now.strftime('%d-%m-%Y')}_{now.strftime('%H-%M-%S')}.csv"
        # Return CSV file
        response = HttpResponse(csv_content, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{default_filename}"'
        
        log_api_response("search_download", request.method, dict(params), {"rows": len(docs)})
        return response

    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("search_download", request.method, dict(request.GET), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

@csrf_exempt
def search_download_sticker(request):
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("search_download_sticker", request.method, dict(request.GET), error_response)
        return JsonResponse(error_response, status=405)

    user, err = require_auth(request)
    if err:
        return err
    
    try:
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        master_file = os.path.join(BASE_DIR, "static", "templates", "Print_Pass_Master_Excel.xlsx")
        wb = load_workbook(master_file)
        ws = wb.active
        params = request.GET
        poffset = int(params.get("offset","0"))
        query = _build_search_query(params)
        docs = list(collection.find(query))
        for doc in docs:
            passNo = doc.get("passNo","")
            dateIn = _format_date_ddmmyyyy(doc.get("dateIn",""))
            items = doc.get("items", [])
            projectName = doc.get("projectName","")
            unitAddress = doc.get("customer",{}).get("unitAddress","")
            for j, item in enumerate(items):
                start_slNo = j+1
                itemName = item.get("itemName", "")
                serialNumber = item.get("serialNumber", "")
                j += poffset-1
                # Choose left or right block based on index
                if j % 2 == 0:
                    col_A, col_B, col_C = "A", "B", "C"
                else:
                    col_A, col_B, col_C = "D", "E", "F"

                start_row = (j // 2) * 7 + 1  # ensures each block gets 7 rows
                if start_row in (57,112):
                    start_row = (j//2) * 7
                    print(start_row)

                labels = ["Pass No", "Date", "Unit Address", "Project Name", "Item Name", "Sl.No"]
                values = [passNo, dateIn, unitAddress, projectName, itemName, serialNumber]

                # write serial number in first column
                ws[f"{col_A}{start_row}"] = start_slNo

                # write labels and values
                for offset, (label, value) in enumerate(zip(labels, values)):
                    ws[f"{col_B}{start_row + offset}"] = label
                    ws[f"{col_C}{start_row + offset}"] = value

                # empty row separator
                empty_row = start_row + len(labels)
                if empty_row != 56:
                    ws[f"{col_B}{empty_row}"] = ""

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        default_filename = f"{datetime.now(ZoneInfo('Asia/Kolkata')).strftime('%Y-%m-%d')}_sticker_export.xlsx"
        response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="{default_filename}"'
        log_api_response("search_download_sticker", request.method, dict(params), {"rows": len(docs)})
        return response
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("search_download_sticker", request.method, dict(request.GET), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

@csrf_exempt
def search_download_form(request):
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        # --- Extract part numbers from query params ---
        part_numbers_str = request.GET.get("PartNumbers", "")
        allowed_part_numbers = set(p.strip() for p in part_numbers_str.split(",") if p.strip())

        query = _build_search_query(request.GET)
        docs = list(collection.find(query))
        if not docs:
            return JsonResponse({"error": "No documents found"}, status=404)

        # ---------- EXCEL SETUP ----------
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            std = wb["Sheet"]
            wb.remove(std)

        # page_number = 1
        # first_doc = docs[0] if docs else {}
        # ws = create_page(wb, page_number, first_doc)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        MAX_ITEMS_PER_PAGE = 10  # Items per page

        # Helper functions
        def style_cell(cell, bold=False, size=10, align="left", wrap=False):
            cell.font = Font(bold=bold, size=size)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

        label_rows = {
            3: [("PASS NO:", "A", "B"), ("DATE:", "E", "E")],
            4: [("PASS DATE:", "A", "B"), ("CUSTOMER NAME:", "E", "E")],
            5: [("PROJECT NAME:", "A", "B"), ("CUSTOMER CONTACT NO:", "E", "E")],
            6: [("UNIT ADDRESS:", "A", "B"), ("CUSTOMER LOCATION:", "E", "E")]
        }

        headers = ["SL. NO", "PART NO", "ITEM NAME", "ITEM S1.N",
                   "DEFECT DETAILS", "RECTIFICATION DETAILS", "RCVD BACK BY CS"]

        column_widths = {'A': 5, 'B': 20, 'C': 25, 'D': 15, 'E': 20, 'F': 40, 'G': 10}

        # ---------- FUNCTIONS ----------
        def create_page(wb, page_number, doc):
            ws = wb.create_sheet(title=f"Customer Support MILCOM - Page {page_number}")
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            for i in range(3, 7):
                ws.row_dimensions[i].height = 10
            ws.row_dimensions[7].height = 5

            # Titles
            ws.merge_cells('A1:G1')
            ws['A1'] = "CUSTOMER SUPPORT MILCOM"
            style_cell(ws['A1'], bold=True, size=16, align="center")

            ws.merge_cells('A2:G2')
            ws['A2'] = "Customer Complaint History Card"
            style_cell(ws['A2'], bold=True, size=12, align="center")

            # Header Labels
            cust = doc.get("customer", {})
            date_in_display = _format_date_ddmmyyyy(doc.get("dateIn", ""))
            print(date_in_display)
            header_values = {
                'C3:D3': doc.get("passNo", ""),
                'F3:G3': datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%d-%m-%Y"),
                'C4:D4': date_in_display,
                'F4:G4': cust.get("name", ""),
                'C5:D5': doc.get("projectName", ""),
                'F5:G5': cust.get("phone", ""),
                'C6:D6': cust.get("unitAddress", ""),
                'F6:G6': cust.get("location", "")
            }

            for row, pairs in label_rows.items():
                for label, start_col, end_col in pairs:
                    if start_col != end_col:
                        ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
                    c = ws[f"{start_col}{row}"]
                    c.value = label
                    style_cell(c, bold=True, size=8, align="left")

            for cells, value in header_values.items():
                ws.merge_cells(cells)
                c = ws[cells.split(":")[0]]
                c.value = value
                style_cell(c, align="left")

            # Table header
            for col, h in enumerate(headers, start=1):
                c = ws.cell(row=8, column=col, value=h)
                style_cell(c, bold=True, size=9, align="center", wrap=True)
                c.border = thin_border

            # --- Set minimum row height for item rows ---
            for r in range(9, 19):
                ws.row_dimensions[r].height = 32

            return ws

        def create_footer(ws, start_row=19, end_row=23):
            footers = ["HANDED OVER BY (CS-Rep)", "RECEIVED BY (TS-Rep)", "REMARKS"]
            ws.merge_cells(f'A{start_row}:B{start_row}')
            ws.merge_cells(f'C{start_row}:D{start_row}')
            ws.merge_cells(f'E{start_row}:G{start_row}')
            ws['A19'], ws['C19'], ws['E19'] = footers
            for cell in ['A19', 'C19', 'E19']:
                style_cell(ws[cell], bold=True)
                ws[cell].alignment = Alignment(horizontal="center", vertical="center")

            for row in range(start_row + 1, end_row + 1):
                for col in range(1, 8):
                    c = ws.cell(row=row, column=col)
                    c.border = thin_border
                    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            ws.merge_cells(f'A{start_row+1}:B{end_row}')
            ws.merge_cells(f'C{start_row+1}:D{end_row}')
            ws.merge_cells(f'E{start_row+1}:G{end_row}')

        # ---------- POPULATE ----------
        serial_number = 1
        current_row = 9
        page_number = 1
        first_doc = docs[0] if docs else {}
        ws = create_page(wb, page_number, first_doc)

        for doc in docs:
            items = [i for i in doc.get("items", []) if str(i.get("partNumber", "")).strip() in allowed_part_numbers]
            for item in items:
                if (current_row - 9) % MAX_ITEMS_PER_PAGE == 0 and current_row != 9:
                    create_footer(ws)
                    page_number += 1
                    ws = create_page(wb, page_number, doc)
                    current_row = 9

                row_values = [
                    serial_number, item.get("partNumber", ""), item.get("itemName", ""),
                    item.get("serialNumber", ""), item.get("defectDetails", ""),
                    "", ""
                ]
                for col, val in enumerate(row_values, start=1):
                    c = ws.cell(row=current_row, column=col, value=val)
                    c.border = thin_border
                    align = "center" if col == 1 else "left"
                    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
                serial_number += 1
                current_row += 1
            if ws:
                create_footer(ws)

        # ---------- RESPONSE ----------
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="Customer_Complaint_History_Card.xlsx"'
        return response

    except Exception as e:
        stack_trace = traceback.format_exc()
        return JsonResponse({"error": str(e), "stack": stack_trace}, status=500)


def search_download_acknowledgement(request):
    """Download acknowledgement form using template with pass no details + user metadata."""
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        pass_no = (request.GET.get("passNo") or "").strip()
        if not pass_no:
            return JsonResponse({"error": "passNo is required"}, status=400)

        doc = collection.find_one({"passNo": pass_no})
        if not doc:
            return JsonResponse({"error": "Pass number not found"}, status=404)

        # Fetch user metadata (name, designation, mobile) from the logged-in user's record
        user_doc = users_collection.find_one({"username": user.get("username")}, {"password_hash": 0})
        user_name = user_doc.get("name", "") if user_doc else ""
        user_designation = user_doc.get("designation", "") if user_doc else ""
        user_mobile = user_doc.get("mobile", "") if user_doc else ""

        # Load template
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        template_path = os.path.join(BASE_DIR, "static", "templates", "Acknowledgement_Receipt_Master.xlsx")
        wb = load_workbook(template_path)
        ws = wb.active

        # Unmerge all merged cells so we can write freely, then re-merge after
        merged_ranges = list(ws.merged_cells.ranges)
        for merge_range in merged_ranges:
            ws.unmerge_cells(str(merge_range))

        customer = doc.get("customer", {})

        # Fill cells per template layout
        ws["D9"] = doc.get("passNo", "")
        ws["D10"] = customer.get("name", "")
        ws["D11"] = customer.get("unitAddress", "")
        raw_date = doc.get("dateIn", "")
        try:
            ws["M9"] = datetime.strptime(raw_date, "%Y-%m-%d").strftime("%d-%m-%Y") if raw_date else ""
        except (ValueError, TypeError):
            ws["M9"] = raw_date
        ws["M10"] = customer.get("phone", "")
        ws["M11"] = customer.get("location", "")
        ws["K30"] = user_name
        ws["K32"] = user_designation
        ws["K34"] = user_mobile

        # Justify alignment for filled cells
        justify_align = Alignment(horizontal="justify", vertical="center", wrap_text=True)
        for cell_ref in ["D9", "D10", "D11", "M9", "M10", "M11", "K30", "K32", "K34"]:
            ws[cell_ref].alignment = justify_align

        # Re-merge cells
        for merge_range in merged_ranges:
            ws.merge_cells(str(merge_range))

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"Acknowledgement_{pass_no}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

    
@csrf_exempt
def search_suggestions(request):
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("search_suggestions", request.method, dict(request.GET), error_response)
        return JsonResponse(error_response, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        params = request.GET
        search_type = params.get("type")
        value = (params.get("value") or "").strip()
        if not search_type or not value:
            response = {"error": "type and value are required"}
            log_api_response("search_suggestions", request.method, dict(params), response)
            return JsonResponse(response, status=400)

        suggestions = set()
        if search_type == "passNo":
            docs = collection.find({"passNo": {"$regex": f"^{re.escape(value)}", "$options": "i"}}, {"_id": 0, "passNo": 1}).limit(10)
            for doc in docs:
                suggestions.add(doc.get("passNo"))
        elif search_type == "ItemPartNo":
            docs = collection.find({"items.partNumber": {"$regex": f"^{re.escape(value)}", "$options": "i"}}, {"_id": 0, "items.partNumber": 1}).limit(50)
            for doc in docs:
                for item in doc.get("items", []):
                    part_no = item.get("partNumber")
                    if part_no and part_no.lower().startswith(value.lower()):
                        suggestions.add(part_no)
                        if len(suggestions) >= 10:
                            break
                if len(suggestions) >= 10:
                    break
        elif search_type == "ProjectName":
            docs = admin_projects_collection.find({"projectName": {"$regex": f"^{re.escape(value)}", "$options": "i"}}, {"_id": 0, "projectName": 1}).limit(10)
            for doc in docs:
                suggestions.add(doc.get("projectName"))
        elif search_type == "PhoneNumber":
            # Show suggestions once we have 3+ digits (prefix match).
            phone_digits = re.sub(r"\D+", "", value)
            if len(phone_digits) < 3:
                suggestions = set()
            else:
                # Try Mongo prefix regex first.
                docs = collection.find(
                    {"customer.phone": {"$regex": f"^{phone_digits}"}},
                    {"_id": 0, "customer.phone": 1}
                ).limit(50)

                for doc in docs:
                    phone = doc.get("customer", {}).get("phone", "")
                    phone_str = str(phone)
                    phone_str_digits = re.sub(r"\D+", "", phone_str)
                    if re.match(rf"^{re.escape(phone_digits)}", phone_str_digits):
                        suggestions.add(phone_str)
                        if len(suggestions) >= 10:
                            break

                # Fallback scan (covers numeric-stored phone fields)
                if len(suggestions) < 1:
                    fallback_docs = collection.find({}, {"_id": 0, "customer.phone": 1}).limit(300)
                    for doc in fallback_docs:
                        phone = doc.get("customer", {}).get("phone", "")
                        phone_str = str(phone)
                        phone_str_digits = re.sub(r"\D+", "", phone_str)
                        if re.match(rf"^{re.escape(phone_digits)}", phone_str_digits):
                            suggestions.add(phone_str)
                            if len(suggestions) >= 10:
                                break
        else:
            response = {"error": "Invalid type"}
            log_api_response("search_suggestions", request.method, dict(params), response)
            return JsonResponse(response, status=400)

        suggestions = list(suggestions)[:10]  # Limit to 10 suggestions
        response = {"suggestions": suggestions}
        log_api_response("search_suggestions", request.method, dict(params), {"count": len(suggestions)})
        return JsonResponse(response)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("search_suggestions", request.method, dict(request.GET), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)


def _normalized_spares_bin_labels(bin_nos):
    out = []
    for b in bin_nos or []:
        s = str(b).strip()
        if s:
            out.append(s)
    return out


def _check_spares_bins_unique_in_store(store, bin_nos, exclude_part_no=None):
    """
    Reject if any bin label is already used by another part in the same store.
    """
    bins = _normalized_spares_bin_labels(bin_nos)
    if len(bins) != len(set(bins)):
        return JsonResponse(
            {"error": "Duplicate bin numbers within this item"},
            status=400,
        )
    if not store or not bins:
        return None
    q = {"item_loc": store}
    if exclude_part_no:
        q["part_no"] = {"$ne": exclude_part_no}
    for doc in spares_master.find(q, {"bin_nos": 1}):
        existing = set(_normalized_spares_bin_labels(doc.get("bin_nos")))
        for b in bins:
            if b in existing:
                msg = (
                    "This bin is already allotted to a different item. "
                    f"Use a different bin. Duplicate Bin: {b}"
                )
                return JsonResponse({"error": msg}, status=409)
    return None


def _spares_master_payload(body, user, *, is_update: bool):
    """Validate and build document fields from JSON body."""
    part_no = (body.get("part_no") or "").strip()
    item_name = (body.get("item_name") or "").strip()
    project_name = (body.get("project_name") or "").strip()
    no_of_bins = body.get("no_of_bins")
    bin_nos = body.get("bin_nos", [])
    rack_no = (body.get("rack_no") or "").strip()
    item_loc = (body.get("item_loc") or "").strip()

    if no_of_bins is None or no_of_bins == "":
        return None, JsonResponse({"error": "no_of_bins is required"}, status=400)
    try:
        no_of_bins_int = int(no_of_bins)
    except (TypeError, ValueError):
        return None, JsonResponse({"error": "no_of_bins must be a number"}, status=400)

    if not isinstance(bin_nos, list) or len(bin_nos) != no_of_bins_int:
        return None, JsonResponse(
            {"error": "bin_nos count must match no_of_bins"},
            status=400,
        )
    if not part_no or not item_name:
        return None, JsonResponse({"error": "part_no and item_name are required"}, status=400)
    if not item_loc:
        return None, JsonResponse({"error": "store (item_loc) is required"}, status=400)

    doc = {
        "part_no": part_no,
        "item_name": item_name,
        "project_name": project_name,
        "no_of_bins": no_of_bins_int,
        "bin_nos": bin_nos,
        "rack_no": rack_no,
        "item_loc": item_loc,
    }
    if not is_update:
        doc["created_by"] = user.get("username")
        doc["created_at"] = datetime.now(ZoneInfo("Asia/Kolkata"))
    else:
        doc["updated_by"] = user.get("username")
        doc["updated_at"] = datetime.now(ZoneInfo("Asia/Kolkata"))
    return doc, None


@csrf_exempt
def spares_master_add(request):
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    user, err = require_auth(request, role="admin")
    if err:
        return err

    try:
        body = json.loads(request.body or "{}")

        doc, bad = _spares_master_payload(body, user, is_update=False)
        if bad:
            return bad

        spares_coll = db["spares_master"]
        part_no = doc["part_no"]

        if spares_coll.find_one({"part_no": part_no}):
            return JsonResponse({"error": "Part number already exists"}, status=409)

        dup_err = _check_spares_bins_unique_in_store(doc["item_loc"], doc["bin_nos"], None)
        if dup_err:
            return dup_err

        spares_coll.insert_one(doc)

        return JsonResponse({"message": "Item added", "part_no": part_no}, status=201)

    except Exception as e:
        stack = traceback.format_exc()
        return JsonResponse({"error": str(e), "stack_trace": stack}, status=500)


@csrf_exempt
def spares_master_update(request):
    """Update existing master row (admin only)."""
    if request.method != "PUT":
        return JsonResponse({"error": "Only PUT allowed"}, status=405)

    user, err = require_auth(request, role="admin")
    if err:
        return err

    try:
        body = json.loads(request.body or "{}")
        doc, bad = _spares_master_payload(body, user, is_update=True)
        if bad:
            return bad

        part_no = doc["part_no"]
        spares_coll = db["spares_master"]
        existing = spares_coll.find_one({"part_no": part_no})
        if not existing:
            return JsonResponse({"error": "Part number not found"}, status=404)

        dup_err = _check_spares_bins_unique_in_store(doc["item_loc"], doc["bin_nos"], part_no)
        if dup_err:
            return dup_err

        # Preserve qty and history if present
        update_fields = {k: v for k, v in doc.items() if k != "part_no"}
        spares_coll.update_one({"part_no": part_no}, {"$set": update_fields})

        return JsonResponse({"message": "Item updated", "part_no": part_no}, status=200)

    except Exception as e:
        stack = traceback.format_exc()
        return JsonResponse({"error": str(e), "stack_trace": stack}, status=500)


@csrf_exempt
def spares_master_delete(request):
    """Delete a spares master item and its related audit/history data (admin only)."""
    if request.method != "DELETE":
        return JsonResponse({"error": "Only DELETE allowed"}, status=405)

    user, err = require_auth(request, role="admin")
    if err:
        return err

    try:
        body = json.loads(request.body or "{}")
        part_no = (body.get("part_no") or "").strip()
        if not part_no:
            return JsonResponse({"error": "part_no is required"}, status=400)

        spares_coll = db["spares_master"]
        existing = spares_coll.find_one({"part_no": part_no})
        if not existing:
            return JsonResponse({"error": "Part number not found"}, status=404)

        # Delete the master record
        spares_coll.delete_one({"part_no": part_no})
        # Delete related audit entries
        spares_audit.delete_many({"part_no": part_no})
        # Delete related spares_in entries
        spares_in_col.delete_many({"part_no": part_no})
        # Delete related spares_out entries
        spares_out_col.delete_many({"part_no": part_no})
        # Delete related returnable entries
        spares_out_returnable_col.delete_many({"part_no": part_no})
        spares_in_returned_col.delete_many({"part_no": part_no})
        spares_returnable_requests_col.delete_many({"part_no": part_no})

        return JsonResponse({"message": "Item and related data deleted", "part_no": part_no})

    except Exception as e:
        stack = traceback.format_exc()
        return JsonResponse({"error": str(e), "stack_trace": stack}, status=500)


@csrf_exempt
def spares_master_search(request):
    """Regex / pattern search on part_no for suggestions (authenticated users)."""
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        pattern = (request.GET.get("pattern") or "").strip()
        if not pattern:
            return JsonResponse({"matches": []})

        # Safe substring search: escape regex metacharacters
        safe = re.escape(pattern)
        cursor = spares_master.find(
            {"part_no": {"$regex": safe, "$options": "i"}},
            {"_id": 0, "part_no": 1, "item_name": 1, "project_name": 1, "item_loc": 1},
        ).limit(25)
        matches = list(cursor)
        return JsonResponse({"matches": matches})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def spares_stores_list(request):
    """Predefined store names for dropdowns (any authenticated user)."""
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        docs = list(spares_stores_col.find({}, {"_id": 0, "name": 1}))
        names = sorted(
            [d.get("name") for d in docs if d.get("name")],
            key=lambda s: str(s).casefold(),
        )
        return JsonResponse({"stores": names})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def admin_stores_add(request):
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    user, err = require_auth(request, role="admin")
    if err:
        return err

    try:
        body = json.loads(request.body or "{}")
        name = (body.get("name") or "").strip()
        if not name:
            return JsonResponse({"error": "Store name is required"}, status=400)
        if spares_stores_col.find_one({"name": name}):
            return JsonResponse({"error": "Store already exists"}, status=409)
        spares_stores_col.insert_one(
            {
                "name": name,
                "createdBy": user.get("username"),
                "createdAt": datetime.now(ZoneInfo("Asia/Kolkata")),
            }
        )
        return JsonResponse({"message": "Store added", "name": name}, status=201)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def admin_stores_edit(request):
    if request.method != "PUT":
        return JsonResponse({"error": "Only PUT allowed"}, status=405)

    user, err = require_auth(request, role="admin")
    if err:
        return err

    try:
        body = json.loads(request.body or "{}")
        old_name = (body.get("oldName") or "").strip()
        new_name = (body.get("newName") or "").strip()
        if not old_name or not new_name:
            return JsonResponse({"error": "oldName and newName are required"}, status=400)
        if old_name == new_name:
            return JsonResponse({"message": "No change", "name": new_name})

        if not spares_stores_col.find_one({"name": old_name}):
            return JsonResponse({"error": "Store not found"}, status=404)
        if spares_stores_col.find_one({"name": new_name}):
            return JsonResponse({"error": "A store with that name already exists"}, status=409)

        spares_stores_col.update_one({"name": old_name}, {"$set": {"name": new_name, "updatedBy": user.get("username"), "updatedAt": datetime.now(ZoneInfo("Asia/Kolkata"))}})
        # Keep master list in sync: item_loc holds selected store name
        spares_master.update_many({"item_loc": old_name}, {"$set": {"item_loc": new_name}})

        return JsonResponse({"message": "Store updated", "name": new_name})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def admin_stores_list(request):
    """List stores (admin) — same data as spares_stores_list but explicit admin route."""
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    user, err = require_auth(request, role="admin")
    if err:
        return err

    try:
        docs = list(spares_stores_col.find({}, {"_id": 0, "name": 1}))
        names = sorted(
            [d.get("name") for d in docs if d.get("name")],
            key=lambda s: str(s).casefold(),
        )
        return JsonResponse({"stores": names})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def spares_master_list(request):
    if request.method == "GET":
        try:
            part_no = request.GET.get("part_no")

            # If part_no is passed → return the single item's details
            if part_no:
                item = spares_master.find_one({"part_no": part_no}, {"_id": 0})
                if not item:
                    return JsonResponse({"error": "Item not found"}, status=404)
                return JsonResponse(item)

            # Otherwise → return full list
            items = list(spares_master.find({}, {"_id": 0}))
            return JsonResponse({"items": items})

        except Exception as e:
            print(traceback.format_exc())
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method Not Allowed"}, status=405)

@csrf_exempt
def spares_in(request):
    user = get_user_from_token(request)
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8"))
            part_no = data.get("part_no")
            recieved_from = data.get("recieved_from")
            qty_in = int(data.get("qty_in", 0))
            remarks = data.get("remarks")

            if not part_no or qty_in <= 0 or not recieved_from:
                return JsonResponse({"error": "Invalid data"}, status=400)

            # Find item in master list
            item = spares_master.find_one({"part_no": part_no})
            if not item:
                return JsonResponse({"error": "Item not found"}, status=404)

            no_of_bins = item.get("no_of_bins", 0)
            bin_nos = item.get("bin_nos", [])
            rack_no = item.get("rack_no", "")
            item_loc = item.get("item_loc", "")
            project_name = item.get("project_name", "")

            # Current qty
            current_qty = int(item.get("qty", 0))

            # Date tracking
            entry_date = datetime.now(ZoneInfo("Asia/Kolkata"))
            date= entry_date

            # New quantity = old + incoming qty
            new_qty = current_qty + qty_in

            # Update master list
            spares_master.update_one(
                {"part_no": part_no},
                {
                    "$set": {"qty": new_qty},
                    "$push": {
                        "history": {
                            "type": "IN",
                            "qty": qty_in,
                            "recieved_from": recieved_from,
                            "date": entry_date,
                            "remarks": remarks,
                            "no_of_bins": no_of_bins,
                            "bin_nos": bin_nos,
                            "rack_no": rack_no,
                            "item_loc": item_loc,
                            "project_name": project_name,
                        }
                    }
                }
            )

            # Log entry 
            spares_in_col.insert_one({
                "part_no": part_no,
                "qty_in": qty_in,
                "recieved_from": recieved_from,
                "previous_qty": current_qty,
                "new_qty": new_qty,
                "date": entry_date,
                "remarks": remarks,
                "no_of_bins": no_of_bins,
                "bin_nos": bin_nos,
                "rack_no": rack_no,
                "item_loc": item_loc,
                "project_name": project_name,
            })

            spares_audit.insert_one({
                "part_no": part_no,
                "date": datetime.now(tz=ZoneInfo("Asia/Kolkata")),
                "recieved_from": recieved_from,
                "in": qty_in,
                "out": 0,
                "qty_after": new_qty,
                "user": user,
                "remarks": remarks,
                "no_of_bins": no_of_bins,
                "bin_nos": bin_nos,
                "rack_no": rack_no,
                "item_loc": item_loc,
                "project_name": project_name,
            })

            return JsonResponse({"status": "success", "new_qty": new_qty})

        except Exception as e:
            print(traceback.format_exc())
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method Not Allowed"}, status=405)

@csrf_exempt
def spares_out(request):
    user = get_user_from_token(request)
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8"))
            part_no = data.get("part_no")
            qty_out = int(data.get("qty_out", 0))
            handing_over_to = data.get("handing_over_to")
            remarks= data.get("remarks")
            if not part_no or qty_out <= 0 or not handing_over_to:
                return JsonResponse({"error": "Invalid input"}, status=400)

            # Find item
            item = spares_master.find_one({"part_no": part_no})
            if not item:
                return JsonResponse({"error": "Item not found"}, status=404)

            no_of_bins = item.get("no_of_bins", 0)
            bin_nos = item.get("bin_nos", [])
            rack_no = item.get("rack_no", "")
            item_loc = item.get("item_loc", "")
            project_name = item.get("project_name", "")

            current_qty = int(item.get("qty", 0))

            # Validate stock availability
            if qty_out > current_qty:
                return JsonResponse({"error": "Not enough stock"}, status=400)

            new_qty = current_qty - qty_out

            # Date
            from datetime import datetime
            from zoneinfo import ZoneInfo
            entry_date = datetime.now(ZoneInfo("Asia/Kolkata"))
            date= entry_date

            # Update master list
            spares_master.update_one(
                {"part_no": part_no},
                {
                    "$set": {"qty": new_qty},
                    "$push": {
                        "history": {
                            "type": "OUT",
                            "qty": qty_out,
                            "handed_to": handing_over_to,
                            "date": entry_date,
                            "remarks": remarks,
                            "no_of_bins": no_of_bins,
                            "bin_nos": bin_nos,
                            "rack_no": rack_no,
                            "item_loc": item_loc,
                            "project_name": project_name,
                        }
                    }
                }
            )

            # Log outgoing
            spares_out_col.insert_one({
                "part_no": part_no,
                "qty_out": qty_out,
                "handing_over_to": handing_over_to,
                "previous_qty": current_qty,
                "new_qty": new_qty,
                "date": entry_date,
                "remarks": remarks,
                "no_of_bins": no_of_bins,
                "bin_nos": bin_nos,
                "rack_no": rack_no,
                "item_loc": item_loc,
                "project_name": project_name,
            })

            spares_audit.insert_one({
                "part_no": part_no,
                "date": datetime.now(tz=ZoneInfo("Asia/Kolkata")),
                "in": 0,
                "out": qty_out,
                "qty_after": new_qty,
                "user": user,
                "remarks": remarks,
                "handing_over_to": handing_over_to,
                "no_of_bins": no_of_bins,
                "bin_nos": bin_nos,
                "rack_no": rack_no,
                "item_loc": item_loc,
                "project_name": project_name,
            })

            return JsonResponse({"status": "success", "new_qty": new_qty})

        except Exception as e:
            print(traceback.format_exc())
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method Not Allowed"}, status=405)

@csrf_exempt
def spares_audit_view(request):
    if request.method == "GET":
        try:
            part_no = request.GET.get("part_no")

            audit_items = list(
                spares_audit.find({"part_no": part_no}, {"_id": 0}).sort("date", 1)
            )

            return JsonResponse({"audit": audit_items})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method Not Allowed"}, status=405)

@csrf_exempt
def spares_audit_filter(request):
    if request.method == "GET":
        try:
            part_no = request.GET.get("part_no")
            start_date = request.GET.get("start_date")
            end_date = request.GET.get("end_date")

            query = {"part_no": part_no}

            # If date filter applied
            if start_date and end_date:
                try:
                    start_dt = datetime.strptime(start_date, "%Y-%m-%d")

                    # end date should include entire day (23:59:59)
                    end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)

                    query["date"] = {
                        "$gte": start_dt,
                        "$lt": end_dt
                    }

                except Exception:
                    print("Date parsing error:", traceback.format_exc())
                    return JsonResponse({"error": "Invalid date format"}, status=400)

            audit_items = list(
                spares_audit.find(query, {"_id": 0}).sort("date", 1)
            )

            return JsonResponse({"audit": audit_items})

        except Exception as e:
            print("Error:", traceback.format_exc())
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method Not Allowed"}, status=405)

def sort_key(part_no):
    part_no = str(part_no).strip()
    match = re.match(r'^(\d+)', part_no)

    if match:
        return (1, int(match.group(1)), part_no)
    else:
        return (2, float('inf'), part_no)

@csrf_exempt
def stock_check(request):
    if request.method == "GET":
        try:
            # Fetch stock same way as stock check
            items = list(spares_master.find({}, {"_id": 0}))
            items.sort(key=lambda x: sort_key(x.get("part_no", "")))

            # Build CSV
            import csv
            from io import StringIO

            output = StringIO()
            writer = csv.writer(output)

            # Header
            writer.writerow(["Sl No", "Part No", "Item Name", "Project Name", "Item Loc", "Rack No", "No of Bins", "Bin No", "Qty"])

            # Rows
            for idx, item in enumerate(items):
                bin_nos_val = item.get("bin_nos")
                if isinstance(bin_nos_val, list):
                    bin_cell = ", ".join(str(b) for b in bin_nos_val)
                else:
                    bin_cell = item.get("bin_no", "") or ""
                def _csv_safe(v):
                    s = str(v) if v is not None else ""
                    return s.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
                writer.writerow([
                    idx + 1,
                    _csv_safe(item.get("part_no", "")),
                    _csv_safe(item.get("item_name", "")),
                    _csv_safe(item.get("project_name", "")),
                    _csv_safe(item.get("item_loc", "")),
                    _csv_safe(item.get("rack_no", "")),
                    item.get("no_of_bins", 0),
                    _csv_safe(bin_cell),
                    item.get("qty", 0)
                ])

            response = HttpResponse(output.getvalue(), content_type="text/csv")
            response["Content-Disposition"] = "attachment; filename=stock_report.csv"
            return response

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method Not Allowed"}, status=405)


# ----------------------
# Spares Returnable Flow
# ----------------------

def _preview_next_service_request_no():
    counters = db["counters"]
    doc = counters.find_one({"_id": "spares_returnable_service_request"}, {"seq": 1})
    current = int((doc or {}).get("seq", 0) or 0)
    return current + 1


def _generate_next_service_request_no():
    counters = db["counters"]
    doc = counters.find_one_and_update(
        {"_id": "spares_returnable_service_request"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    return int(doc.get("seq", 1))


def spares_returnable_next_service_request(request):
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        return JsonResponse({"nextServiceRequestNo": _preview_next_service_request_no()})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def spares_out_returnable(request):
    user = get_user_from_token(request)
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body.decode("utf-8"))
        part_no = (data.get("part_no") or "").strip()
        qty_out = int(data.get("qty_out", 0))
        date_out = (data.get("date") or datetime.now(ZoneInfo("Asia/Kolkata")).date().isoformat()).strip()
        handed_over_by_cs = (data.get("handed_over_by_cs") or "").strip()
        received_by_ts = (data.get("received_by_ts") or "").strip()
        remarks = (data.get("remarks") or "").strip()

        if not part_no or qty_out <= 0 or not handed_over_by_cs or not received_by_ts:
            return JsonResponse({"error": "Invalid input"}, status=400)

        item = spares_master.find_one({"part_no": part_no})
        if not item:
            return JsonResponse({"error": "Item not found"}, status=404)

        current_qty = int(item.get("qty", 0))
        if qty_out > current_qty:
            return JsonResponse({"error": "Not enough stock"}, status=400)

        new_qty = current_qty - qty_out
        service_request_no = _generate_next_service_request_no()
        entry_date = datetime.now(ZoneInfo("Asia/Kolkata"))

        spares_master.update_one(
            {"part_no": part_no},
            {
                "$set": {"qty": new_qty},
                "$push": {
                    "history": {
                        "type": "OUT_RETURNABLE",
                        "service_request_no": service_request_no,
                        "qty": qty_out,
                        "date": entry_date,
                        "remarks": remarks,
                        "handed_over_by_cs": handed_over_by_cs,
                        "received_by_ts": received_by_ts,
                        "project_name": item.get("project_name", ""),
                        "item_name": item.get("item_name", ""),
                        "item_loc": item.get("item_loc", ""),
                        "rack_no": item.get("rack_no", ""),
                    }
                },
            },
        )

        request_doc = {
            "serviceRequestNo": service_request_no,
            "date": date_out,
            "part_no": part_no,
            "item_name": item.get("item_name", ""),
            "project_name": item.get("project_name", ""),
            "qty_handed_over": qty_out,
            "handed_over_by_cs": handed_over_by_cs,
            "received_by_ts": received_by_ts,
            "remarks": remarks,
            "returns": [],
            "createdBy": (user or {}).get("username"),
            "createdAt": entry_date,
            "updatedAt": entry_date,
            "status": "OPEN",
        }
        spares_returnable_requests_col.insert_one(request_doc)

        spares_out_returnable_col.insert_one(
            {
                "serviceRequestNo": service_request_no,
                "part_no": part_no,
                "qty_out": qty_out,
                "previous_qty": current_qty,
                "new_qty": new_qty,
                "date": entry_date,
                "remarks": remarks,
                "handed_over_by_cs": handed_over_by_cs,
                "received_by_ts": received_by_ts,
                "project_name": item.get("project_name", ""),
                "item_name": item.get("item_name", ""),
                "item_loc": item.get("item_loc", ""),
                "rack_no": item.get("rack_no", ""),
            }
        )

        spares_audit.insert_one(
            {
                "part_no": part_no,
                "date": entry_date,
                "service_request_no": service_request_no,
                "in": 0,
                "out": qty_out,
                "qty_after": new_qty,
                "user": user,
                "remarks": remarks,
                "project_name": item.get("project_name", ""),
                "returnable": True,
                "handed_over_by_cs": handed_over_by_cs,
                "received_by_ts": received_by_ts,
            }
        )

        return JsonResponse(
            {
                "status": "success",
                "serviceRequestNo": service_request_no,
                "new_qty": new_qty,
            }
        )
    except Exception as e:
        print(traceback.format_exc())
        return JsonResponse({"error": str(e)}, status=500)


def _build_returnable_summary(doc):
    returns = doc.get("returns", []) or []
    total_returned = sum(int(r.get("qtyReturned", 0) or 0) for r in returns)
    qty_handed = int(doc.get("qty_handed_over", 0) or 0)
    outstanding = max(qty_handed - total_returned, 0)
    return {
        "serviceRequestNo": doc.get("serviceRequestNo"),
        "date": doc.get("date", ""),
        "part_no": doc.get("part_no", ""),
        "item_name": doc.get("item_name", ""),
        "project_name": doc.get("project_name", ""),
        "qty_handed_over": qty_handed,
        "qty_returned": total_returned,
        "outstanding_qty": outstanding,
        "status": "CLOSED" if outstanding == 0 else "OPEN",
    }


def spares_out_returnable_list(request):
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        docs = list(
            spares_returnable_requests_col.find({}, {"_id": 0}).sort([("serviceRequestNo", -1)]).limit(500)
        )
        summaries = [_build_returnable_summary(d) for d in docs]
        return JsonResponse({"requests": summaries})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def spares_out_returnable_record(request, service_request_no):
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        doc = spares_returnable_requests_col.find_one({"serviceRequestNo": int(service_request_no)}, {"_id": 0})
        if not doc:
            return JsonResponse({"error": "Service Request not found"}, status=404)

        summary = _build_returnable_summary(doc)
        response = {**doc, **summary}
        return JsonResponse(response)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def spares_in_returned(request):
    user = get_user_from_token(request)
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body.decode("utf-8"))
        service_request_no_raw = data.get("serviceRequestNo")
        qty_in = int(data.get("qty_in", 0))
        date_in = (data.get("date_in") or datetime.now(ZoneInfo("Asia/Kolkata")).date().isoformat()).strip()
        received_from = (data.get("received_from") or "").strip()
        received_back_by_cs = (data.get("received_back_by_cs") or (user or {}).get("username") or "").strip()
        remarks = (data.get("remarks") or "").strip()

        if service_request_no_raw in (None, ""):
            return JsonResponse({"error": "Service Request No is required"}, status=400)
        service_request_no = int(service_request_no_raw)
        if qty_in <= 0 or not received_from:
            return JsonResponse({"error": "Invalid input"}, status=400)

        req_doc = spares_returnable_requests_col.find_one({"serviceRequestNo": service_request_no})
        if not req_doc:
            return JsonResponse({"error": "Service Request not found"}, status=404)

        returns = req_doc.get("returns", []) or []
        qty_handed = int(req_doc.get("qty_handed_over", 0) or 0)
        total_returned = sum(int(r.get("qtyReturned", 0) or 0) for r in returns)
        outstanding = max(qty_handed - total_returned, 0)
        if qty_in > outstanding:
            return JsonResponse({"error": f"Qty In exceeds outstanding qty ({outstanding})"}, status=400)

        part_no = req_doc.get("part_no")
        item = spares_master.find_one({"part_no": part_no})
        if not item:
            return JsonResponse({"error": "Item not found in master"}, status=404)

        current_qty = int(item.get("qty", 0))
        new_qty = current_qty + qty_in
        entry_date = datetime.now(ZoneInfo("Asia/Kolkata"))

        spares_master.update_one(
            {"part_no": part_no},
            {
                "$set": {"qty": new_qty},
                "$push": {
                    "history": {
                        "type": "IN_RETURNED",
                        "service_request_no": service_request_no,
                        "qty": qty_in,
                        "date": entry_date,
                        "received_from": received_from,
                        "received_back_by_cs": received_back_by_cs,
                        "remarks": remarks,
                        "project_name": item.get("project_name", ""),
                        "item_name": item.get("item_name", ""),
                        "item_loc": item.get("item_loc", ""),
                        "rack_no": item.get("rack_no", ""),
                    }
                },
            },
        )

        next_sl = len(returns) + 1
        return_row = {
            "slNo": next_sl,
            "qtyReturnDate": date_in,
            "qtyReturned": qty_in,
            "handedOverByTs": received_from,
            "receivedBackByCs": received_back_by_cs,
        }

        remaining_after = max(outstanding - qty_in, 0)
        spares_returnable_requests_col.update_one(
            {"serviceRequestNo": service_request_no},
            {
                "$push": {"returns": return_row},
                "$set": {
                    "updatedAt": entry_date,
                    "status": "CLOSED" if remaining_after == 0 else "OPEN",
                },
            },
        )

        spares_in_returned_col.insert_one(
            {
                "serviceRequestNo": service_request_no,
                "part_no": part_no,
                "qty_in": qty_in,
                "previous_qty": current_qty,
                "new_qty": new_qty,
                "date": entry_date,
                "date_in": date_in,
                "received_from": received_from,
                "received_back_by_cs": received_back_by_cs,
                "remarks": remarks,
            }
        )

        spares_audit.insert_one(
            {
                "part_no": part_no,
                "date": entry_date,
                "service_request_no": service_request_no,
                "in": qty_in,
                "out": 0,
                "qty_after": new_qty,
                "user": user,
                "remarks": remarks,
                "project_name": item.get("project_name", ""),
                "returnable": True,
                "received_from": received_from,
                "received_back_by_cs": received_back_by_cs,
            }
        )

        return JsonResponse(
            {
                "status": "success",
                "new_qty": new_qty,
                "outstanding_qty": remaining_after,
            }
        )
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ─────────────── Configuration Management ───────────────

@csrf_exempt
def config_add(request):
    """Add or update a configuration detail entry (upsert by combination)"""
    user, err = require_auth(request)
    if err:
        return err
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)
    try:
        body = json.loads(request.body)
        project_name = (body.get("project_name") or "").strip()
        item_type = (body.get("item_type") or "").strip()
        item_name = (body.get("item_name") or "").strip()
        part_no = (body.get("part_no") or "").strip()
        config_details = (body.get("config_details") or "").strip()

        if not project_name or not item_name or not part_no or not config_details:
            return JsonResponse({"error": "project_name, item_name, part_no, and config_details are required"}, status=400)

        # Check if combination already exists
        filter_q = {
            "project_name": project_name,
            "item_type": item_type,
            "item_name": item_name,
            "part_no": part_no,
        }
        existing = config_details_col.find_one(filter_q)

        if existing:
            # Update existing record
            config_details_col.update_one(filter_q, {"$set": {
                "config_details": config_details,
                "updated_by": user.get("username", ""),
                "updated_at": datetime.now(ZoneInfo("Asia/Kolkata")).isoformat(),
            }})
            return JsonResponse({"status": "success", "message": "Configuration updated"})
        else:
            # Insert new record
            doc = {
                "project_name": project_name,
                "item_type": item_type,
                "item_name": item_name,
                "part_no": part_no,
                "config_details": config_details,
                "created_by": user.get("username", ""),
                "created_at": datetime.now(ZoneInfo("Asia/Kolkata")).isoformat(),
            }
            config_details_col.insert_one(doc)
            return JsonResponse({"status": "success", "message": "Configuration added"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def config_get(request):
    """Get a single configuration record by exact combination"""
    user, err = require_auth(request)
    if err:
        return err
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)
    try:
        project_name = request.GET.get("project_name", "").strip()
        item_type = request.GET.get("item_type", "").strip()
        item_name = request.GET.get("item_name", "").strip()
        part_no = request.GET.get("part_no", "").strip()

        if not project_name or not item_name or not part_no:
            return JsonResponse({"error": "project_name, item_name, and part_no are required"}, status=400)

        doc = config_details_col.find_one({
            "project_name": project_name,
            "item_type": item_type,
            "item_name": item_name,
            "part_no": part_no,
        }, {"_id": 0})

        if doc:
            return JsonResponse({"found": True, "record": doc})
        else:
            return JsonResponse({"found": False})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def config_list(request):
    """List configuration details with optional filters"""
    user, err = require_auth(request)
    if err:
        return err
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)
    try:
        query = {}
        project_name = request.GET.get("project_name", "").strip()
        part_no = request.GET.get("part_no", "").strip()
        if project_name:
            query["project_name"] = project_name
        if part_no:
            query["part_no"] = re.compile("^" + re.escape(part_no) + "$", re.IGNORECASE)

        docs = list(config_details_col.find(query, {"_id": 0}).sort("created_at", -1))
        return JsonResponse({"records": docs})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    except Exception as e:
        print(traceback.format_exc())
        return JsonResponse({"error": str(e)}, status=500)


def spares_out_returnable_download_form(request):
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        sr_raw = request.GET.get("serviceRequestNo")
        if sr_raw in (None, ""):
            return JsonResponse({"error": "serviceRequestNo is required"}, status=400)

        service_request_no = int(sr_raw)
        doc = spares_returnable_requests_col.find_one({"serviceRequestNo": service_request_no}, {"_id": 0})
        if not doc:
            return JsonResponse({"error": "Service Request not found"}, status=404)

            # Load template
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        template_path = os.path.join(BASE_DIR, "static", "templates", "Service_Master_Request_Form.xlsx")
        wb = load_workbook(template_path)
        ws = wb.active

        # Fill cells per template layout
        ws["D8"] = doc.get("serviceRequestNo", "")
        ws["D10"] = doc.get("project_name", "")
        ws["D13"] = doc.get("item_name", "")
        raw_date = doc.get("date", "")
        try:
            ws["M8"] = datetime.strptime(raw_date, "%Y-%m-%d").strftime("%d-%m-%Y") if raw_date else ""
        except (ValueError, TypeError):
            ws["M8"] = raw_date
        ws["M13"] = doc.get("part_no", "")
        ws["D15"] = doc.get("remarks", "")
        ws["C18"] = doc.get("qty_handed_over", "")
        ws["F18"] = doc.get("handed_over_by_cs", "")
        ws["J18"] = doc.get("received_by_ts", "")

        # Justify alignment for filled cells
        justify_align = Alignment(horizontal="justify", vertical="center", wrap_text=True)
        for cell_ref in ["D8", "D10", "D13", "M8", "M13", "D15", "C18", "F18", "J18"]:
            ws[cell_ref].alignment = justify_align

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"Service_Request_{service_request_no}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ----------------------
# OBD Management
# ----------------------

def obd_suggestions(request):
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("obd_suggestions", request.method, dict(request.GET), error_response)
        return JsonResponse(error_response, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        value = (request.GET.get("value") or "").strip()
        # Start suggesting only when first 3 digits are entered.
        if len(value) < 3:
            return JsonResponse({"suggestions": []})

        docs = obd_collection.find({}, {"_id": 0, "obdNo": 1}).sort("obdNo", 1).limit(3000)
        suggestions = []
        for doc in docs:
            obd_val = doc.get("obdNo")
            if obd_val is None:
                continue
            obd_str = str(obd_val)
            if obd_str.startswith(value):
                suggestions.append(obd_str)
            if len(suggestions) >= 10:
                break
        response = {"suggestions": suggestions}
        log_api_response("obd_suggestions", request.method, dict(request.GET), {"count": len(suggestions)})
        return JsonResponse(response)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("obd_suggestions", request.method, dict(request.GET), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

@csrf_exempt
def obd_out(request):
    if request.method != "POST":
        error_response = {"error": "Only POST allowed"}
        log_api_response("obd_out", request.method, getattr(request, "body", None), error_response)
        return JsonResponse(error_response, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        body = json.loads(request.body or b"{}")
        obd_no_raw = body.get("obdNo")
        date_val = (body.get("date") or datetime.now(ZoneInfo("Asia/Kolkata")).date().isoformat()).strip()
        sent_to_location = (body.get("sentToLocation") or "").strip()
        authorized_by = (body.get("authorizedBy") or "").strip()
        project_name = (body.get("projectName") or "").strip()
        item_details = (body.get("itemDetails") or "").strip()
        courier_name = (body.get("courierName") or "").strip()
        docket_number = (body.get("docketNumber") or "").strip()
        docket_status = (body.get("docketStatus") or "In Transit").strip()
        delivered_date = (body.get("deliveredDate") or "").strip()

        if obd_no_raw in (None, ""):
            return JsonResponse({"error": "OBD Number is required"}, status=400)
        try:
            obd_no = int(obd_no_raw)
        except (TypeError, ValueError):
            return JsonResponse({"error": "OBD Number must be an integer"}, status=400)

        if obd_no < 0:
            return JsonResponse({"error": "OBD Number must be non-negative"}, status=400)

        if not date_val:
            return JsonResponse({"error": "Date is required"}, status=400)

        if not sent_to_location:
            return JsonResponse({"error": "Sent To Location is required"}, status=400)

        if not authorized_by:
            return JsonResponse({"error": "Authorised By is required"}, status=400)

        if not project_name:
            return JsonResponse({"error": "Project is required"}, status=400)

        if not item_details:
            return JsonResponse({"error": "Item Details is required"}, status=400)

        if docket_status not in ("Delivered", "In Transit"):
            return JsonResponse({"error": "Docket Status must be Delivered or In Transit"}, status=400)

        if docket_status == "Delivered" and not delivered_date:
            return JsonResponse({"error": "Delivered Date is required when Docket Status is Delivered"}, status=400)

        if docket_status == "In Transit":
            delivered_date = ""

        if obd_collection.find_one({"obdNo": obd_no}):
            return JsonResponse({"error": "OBD Number already exists"}, status=409)

        doc = {
            "obdNo": obd_no,
            "date": date_val,
            "sentToLocation": sent_to_location,
            "authorizedBy": authorized_by,
            "projectName": project_name,
            "itemDetails": item_details,
            "courierName": courier_name,
            "docketNumber": docket_number,
            "docketStatus": docket_status,
            "deliveredDate": delivered_date,
            "createdBy": user.get("username"),
            "createdAt": datetime.now(ZoneInfo("Asia/Kolkata")),
            "updatedBy": user.get("username"),
            "updatedAt": datetime.now(ZoneInfo("Asia/Kolkata")),
        }
        obd_collection.insert_one(doc)
        doc.pop("_id", None)

        response = {"message": "OBD Out recorded", "data": doc}
        log_api_response("obd_out", request.method, {"obdNo": obd_no}, response)
        return JsonResponse(response, status=201)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("obd_out", request.method, getattr(request, "body", None), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)


@csrf_exempt
def obd_record(request, obd_no):
    user, err = require_auth(request)
    if err:
        return err

    try:
        if request.method == "GET":
            doc = obd_collection.find_one({"obdNo": int(obd_no)}, {"_id": 0})
            if not doc:
                response = {"error": "Not found"}
                log_api_response("obd_record", request.method, {"obdNo": obd_no}, response)
                return JsonResponse(response, status=404)
            log_api_response("obd_record", request.method, {"obdNo": obd_no}, {"found": True})
            return JsonResponse(doc, safe=False)

        if request.method == "PUT":
            body = json.loads(request.body or b"{}")
            new_obd_raw = body.get("obdNo", obd_no)
            try:
                new_obd_no = int(new_obd_raw)
            except (TypeError, ValueError):
                return JsonResponse({"error": "OBD Number must be an integer"}, status=400)

            if new_obd_no != int(obd_no):
                if obd_collection.find_one({"obdNo": new_obd_no}):
                    return JsonResponse({"error": "OBD Number already exists"}, status=409)

            set_fields = {
                "obdNo": new_obd_no,
                "date": (body.get("date") or "").strip(),
                "projectName": (body.get("projectName") or "").strip(),
                "itemDetails": (body.get("itemDetails") or "").strip(),
                "sentToLocation": (body.get("sentToLocation") or "").strip(),
                "authorizedBy": (body.get("authorizedBy") or "").strip(),
                "courierName": (body.get("courierName") or "").strip(),
                "docketNumber": (body.get("docketNumber") or "").strip(),
                "docketStatus": (body.get("docketStatus") or "In Transit").strip(),
                "deliveredDate": (body.get("deliveredDate") or "").strip(),
                "updatedBy": user.get("username"),
                "updatedAt": datetime.now(ZoneInfo("Asia/Kolkata")),
            }

            if not set_fields["date"]:
                return JsonResponse({"error": "Date is required"}, status=400)
            if not set_fields["projectName"]:
                return JsonResponse({"error": "Project is required"}, status=400)
            if set_fields["docketStatus"] not in ("Delivered", "In Transit"):
                return JsonResponse({"error": "Docket Status must be Delivered or In Transit"}, status=400)
            if set_fields["docketStatus"] == "Delivered" and not set_fields["deliveredDate"]:
                return JsonResponse({"error": "Delivered Date is required when Docket Status is Delivered"}, status=400)
            if set_fields["docketStatus"] == "In Transit":
                set_fields["deliveredDate"] = ""

            result = obd_collection.update_one({"obdNo": int(obd_no)}, {"$set": set_fields})
            if result.matched_count == 0:
                response = {"error": "Not found"}
                log_api_response("obd_record", request.method, {"obdNo": obd_no}, response)
                return JsonResponse(response, status=404)

            response = {"message": "OBD record updated"}
            log_api_response("obd_record", request.method, {"obdNo": obd_no}, response)
            return JsonResponse(response)

        error_response = {"error": "Method not allowed"}
        log_api_response("obd_record", request.method, {"obdNo": obd_no}, error_response)
        return JsonResponse(error_response, status=405)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("obd_record", request.method, {"obdNo": obd_no}, {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)


def obd_status(request):
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("obd_status", request.method, dict(request.GET), error_response)
        return JsonResponse(error_response, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        params = request.GET
        query = _build_obd_status_query(params)
        docs = list(obd_collection.find(query, {"_id": 0}).sort([("date", -1), ("obdNo", -1)]))
        response = {"count": len(docs), "data": docs}
        log_api_response("obd_status", request.method, dict(params), {"count": len(docs)})
        return JsonResponse(response)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("obd_status", request.method, dict(request.GET), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)


def _build_obd_status_query(params):
    from_date = (params.get("from") or "").strip()
    to_date = (params.get("to") or "").strip()
    docket_status = (params.get("docketStatus") or "All").strip()

    query = {}
    date_cond = _build_date_filter(from_date, to_date)
    if date_cond:
        query["date"] = date_cond

    if docket_status == "Present":
        query["docketNumber"] = {"$nin": ["", None]}
    elif docket_status == "Absent":
        query["$or"] = [
            {"docketNumber": ""},
            {"docketNumber": None},
            {"docketNumber": {"$exists": False}},
        ]
    return query


def obd_status_download(request):
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("obd_status_download", request.method, dict(request.GET), error_response)
        return JsonResponse(error_response, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        params = request.GET
        query = _build_obd_status_query(params)
        docs = list(obd_collection.find(query, {"_id": 0}).sort([("date", -1), ("obdNo", -1)]))

        wb = Workbook()
        ws = wb.active
        ws.title = "OBD Status"

        headers = [
            "SL NO",
            "OBD NO",
            "DATE",
            "SENT TO LOCATION",
            "AUTHORISED BY",
            "PROJECT",
            "ITEM DETAILS",
            "COURIER NAME",
            "DOCKET NUMBER",
            "DOCKET STATUS",
            "DELIVERY STATUS",
            "DELIVERED DATE",
        ]
        ws.append(headers)

        header_font = Font(bold=True)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for idx, row in enumerate(docs, start=1):
            docket_number = row.get("docketNumber") or ""
            docket_presence = "Present" if docket_number else "Absent"
            ws.append([
                idx,
                row.get("obdNo", ""),
                row.get("date", ""),
                row.get("sentToLocation", ""),
                row.get("authorizedBy", ""),
                row.get("projectName", ""),
                row.get("itemDetails", ""),
                row.get("courierName", ""),
                docket_number,
                docket_presence,
                row.get("docketStatus", ""),
                row.get("deliveredDate", ""),
            ])

        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for c in r:
                c.border = border
                c.alignment = Alignment(vertical="top", wrap_text=True)

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 45)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        now = datetime.now(ZoneInfo("Asia/Kolkata"))
        filename = f"OBD_{now.strftime('%d-%m-%Y_%H-%M-%S')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        log_api_response("obd_status_download", request.method, dict(params), {"count": len(docs)})
        return response
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("obd_status_download", request.method, dict(request.GET), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)
